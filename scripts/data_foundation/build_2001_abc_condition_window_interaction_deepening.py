#!/usr/bin/env python3
"""Build condition-window interaction deepening artifacts for 2001 ABC.

This script uses coded structured values only. It does not export raw narrative
text, row ids, or source record content. The output is an unreviewed Falcon Lab
evidence layer for SCIMA/FCHMA interaction analysis, not a condition-to-support
lookup table and not a support adequacy judgment.
"""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
DATASET_ID = "2001_ABC_survey"
SOURCE_DIR = REPO_ROOT / "data/original_secure/structured/2001_ABC_survey"
DERIVED_DIR = REPO_ROOT / "references/derived/scima-fchma/2001-abc-survey-v0-2026-05-22"

OUTPUT_JSON = DERIVED_DIR / "2001-abc-survey-condition-window-interaction-deepening-v0-2026-05-23.json"
OUTPUT_MD = DERIVED_DIR / "2001-abc-survey-condition-window-interaction-deepening-v0-2026-05-23.md"
CHAT_CARDS_JSONL = DERIVED_DIR / "2001-abc-survey-condition-window-falcon-chat-cards-v0-2026-05-23.jsonl"


B_GROUP_COLUMNS = {
    "training_skill_education": list(range(26, 30)),
    "job_task_design_and_performance": list(range(30, 36)),
    "stress_responsibility_safety": list(range(36, 41)),
    "information_communication": list(range(41, 46)),
    "within_workplace_mobility": list(range(46, 53)),
    "commuting": list(range(53, 62)),
    "health_time_self_care_environment": list(range(62, 70)),
    "interpersonal_relations_and_workplace_culture": list(range(70, 76)),
    "emergency_disaster": list(range(76, 79)),
    "offwork_daily_living": list(range(79, 83)),
}

C_GROUP_COLUMNS = {
    "training_skill_education": list(range(44, 48)),
    "job_task_design_and_performance": list(range(48, 54)),
    "stress_responsibility_safety": list(range(54, 59)),
    "information_communication": list(range(59, 64)),
    "within_workplace_mobility": list(range(64, 71)),
    "commuting": list(range(71, 80)),
    "health_time_self_care_environment": list(range(80, 88)),
    "interpersonal_relations_and_workplace_culture": list(range(88, 94)),
    "emergency_disaster": list(range(94, 97)),
    "offwork_daily_living": list(range(97, 102)),
}

SUPPORT_GROUP_ROUTES = {
    "training_skill_education": ["QR-05-entry-prework-translation", "QR-07-quality-career-and-value-translation"],
    "job_task_design_and_performance": ["QR-03-worksite-contact-and-mobility", "QR-07-quality-career-and-value-translation"],
    "stress_responsibility_safety": ["QR-01-health-time-work-design", "QR-03-worksite-contact-and-mobility", "QR-07-quality-career-and-value-translation"],
    "information_communication": ["QR-02-information-work-procedure", "QR-06-disclosure-boundary-and-mutual-translation"],
    "within_workplace_mobility": ["QR-03-worksite-contact-and-mobility"],
    "commuting": ["QR-03-worksite-contact-and-mobility", "QR-04-life-security-sequencing"],
    "health_time_self_care_environment": ["QR-01-health-time-work-design", "QR-04-life-security-sequencing"],
    "interpersonal_relations_and_workplace_culture": ["QR-06-disclosure-boundary-and-mutual-translation", "QR-07-quality-career-and-value-translation"],
    "emergency_disaster": ["QR-03-worksite-contact-and-mobility", "QR-08-diversity-conditioned-same-structure"],
    "offwork_daily_living": ["QR-04-life-security-sequencing", "QR-06-disclosure-boundary-and-mutual-translation"],
}

SUPPORT_GROUP_FREEDOM = {
    "training_skill_education": "learning_variability_freedom",
    "job_task_design_and_performance": "work_content_freedom",
    "stress_responsibility_safety": "load_quality_safety_freedom",
    "information_communication": "information_translation_freedom",
    "within_workplace_mobility": "body_environment_freedom",
    "commuting": "life_commuting_freedom",
    "health_time_self_care_environment": "health_time_freedom",
    "interpersonal_relations_and_workplace_culture": "social_participation_freedom",
    "emergency_disaster": "safety_contingency_freedom",
    "offwork_daily_living": "life_security_freedom",
}

SUPPORT_GROUP_LABELS = {
    "training_skill_education": "training / skill / education translation",
    "job_task_design_and_performance": "job-task design and performance translation",
    "stress_responsibility_safety": "load, responsibility, stress, and safety design",
    "information_communication": "information, communication, and procedure translation",
    "within_workplace_mobility": "within-workplace mobility and body-environment fit",
    "commuting": "commuting and work-life bridge",
    "health_time_self_care_environment": "health time, self-care, and environment",
    "interpersonal_relations_and_workplace_culture": "interpersonal relation and workplace culture",
    "emergency_disaster": "emergency, disaster, and safety contingency",
    "offwork_daily_living": "off-work daily living and life-security bridge",
}

CONTACT_BLOCKS = {
    "C-1": {
        "columns": list(range(86, 101)),
        "label": "information access and presentation contact",
        "routes": ["QR-02-information-work-procedure", "QR-03-worksite-contact-and-mobility"],
        "freedom_axis": "information_access_freedom",
    },
    "C-2": {
        "columns": list(range(101, 113)),
        "label": "judgment, processing, memory, and measurement contact",
        "routes": ["QR-03-worksite-contact-and-mobility", "QR-07-quality-career-and-value-translation"],
        "freedom_axis": "learning_variability_freedom",
    },
    "C-3": {
        "columns": list(range(113, 124)),
        "label": "operation, grip, layout, and physical contact",
        "routes": ["QR-03-worksite-contact-and-mobility"],
        "freedom_axis": "body_environment_freedom",
    },
    "C-4": {
        "columns": list(range(124, 133)),
        "label": "response, warning, danger, evacuation, and document contact",
        "routes": ["QR-03-worksite-contact-and-mobility", "QR-02-information-work-procedure", "QR-08-diversity-conditioned-same-structure"],
        "freedom_axis": "safety_contingency_freedom",
    },
}

CONSTRAINTS = {
    "understanding_knowledge_application": {"column": 134, "routes": ["QR-07-quality-career-and-value-translation", "QR-02-information-work-procedure"], "freedom_axis": "learning_variability_freedom"},
    "productivity_in_task_execution": {"column": 135, "routes": ["QR-07-quality-career-and-value-translation", "QR-03-worksite-contact-and-mobility"], "freedom_axis": "evaluation_participation_freedom"},
    "regular_attendance_time_reliability": {"column": 136, "routes": ["QR-01-health-time-work-design", "QR-07-quality-career-and-value-translation"], "freedom_axis": "health_time_freedom"},
    "posture_and_position_change": {"column": 137, "routes": ["QR-03-worksite-contact-and-mobility", "QR-01-health-time-work-design"], "freedom_axis": "body_environment_freedom"},
    "machine_operation_object_handling": {"column": 138, "routes": ["QR-03-worksite-contact-and-mobility"], "freedom_axis": "operation_freedom"},
    "stress_responsibility_coping": {"column": 139, "routes": ["QR-01-health-time-work-design", "QR-07-quality-career-and-value-translation"], "freedom_axis": "load_quality_safety_freedom"},
    "communication_information_exchange": {"column": 140, "routes": ["QR-02-information-work-procedure"], "freedom_axis": "information_translation_freedom"},
    "within_workplace_mobility": {"column": 141, "routes": ["QR-03-worksite-contact-and-mobility"], "freedom_axis": "body_environment_freedom"},
    "commuting_transport_use": {"column": 142, "routes": ["QR-03-worksite-contact-and-mobility", "QR-04-life-security-sequencing"], "freedom_axis": "life_commuting_freedom"},
    "health_self_care_daily_management": {"column": 143, "routes": ["QR-01-health-time-work-design", "QR-04-life-security-sequencing"], "freedom_axis": "health_time_freedom"},
    "interpersonal_relations_at_work": {"column": 144, "routes": ["QR-06-disclosure-boundary-and-mutual-translation", "QR-07-quality-career-and-value-translation"], "freedom_axis": "social_participation_freedom"},
    "offwork_daily_living_independence": {"column": 145, "routes": ["QR-04-life-security-sequencing"], "freedom_axis": "life_security_freedom"},
}

CONDITION_WINDOWS = {
    "primary_disability": {"column": 16, "label": "主障害", "min_n": 50},
    "severity": {"column": 20, "label": "障害程度", "min_n": 80},
    "onset_period": {"column": 38, "label": "障害をもつようになった時期", "min_n": 80},
    "response_method": {"column": 4, "label": "回答方法", "min_n": 80},
    "visual_function": {"column": 22, "label": "視覚障害", "min_n": 30},
    "hearing_function": {"column": 25, "label": "聴覚障害", "min_n": 50},
    "upper_lower_limb_side": {"column": 27, "label": "上肢・下肢障害（片側・両側）", "min_n": 80},
    "upper_lower_limb_function": {"column": 28, "label": "上肢・下肢障害（欠損・機能全廃・機能障害）", "min_n": 80},
    "kidney_function": {"column": 32, "label": "じん臓機能障害", "min_n": 30},
    "bladder_rectal_function": {"column": 34, "label": "ぼうこう・直腸機能障害", "min_n": 20},
    "other_condition_marker": {"column": 36, "label": "その他の障害", "min_n": 20},
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def numeric(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def rate(num: int, den: int) -> float | None:
    return round(num / den, 4) if den else None


def delta(value: float | None, base: float | None) -> float | None:
    if value is None or base is None:
        return None
    return round(value - base, 4)


def parse_value_labels(label: str) -> dict[int, str]:
    labels: dict[int, str] = {}
    for part in re.split(r",\s*", clean(label)):
        if "=" not in part:
            continue
        key, value = part.split("=", 1)
        key = key.strip()
        if key.isdigit():
            labels[int(key)] = value.strip()
    return labels


def dictionary(workbook: str) -> dict[int, dict[str, Any]]:
    wb = load_workbook(SOURCE_DIR / workbook, read_only=True, data_only=True)
    ws = wb["データ一覧"]
    rows: dict[int, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        idx, variable, question, content, _desc, _num, label = row[:7]
        if isinstance(idx, int):
            rows[idx] = {
                "column": idx,
                "variable": clean(variable),
                "question": clean(question),
                "content": clean(content),
                "label": clean(label),
                "value_labels": parse_value_labels(clean(label)),
            }
    wb.close()
    return rows


def load_rows(workbook: str, sheet: str) -> list[tuple[Any, ...]]:
    wb = load_workbook(SOURCE_DIR / workbook, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    return rows


def val(row: tuple[Any, ...], col: int) -> int | None:
    if col - 1 >= len(row):
        return None
    return numeric(row[col - 1])


def has_any_response(row: tuple[Any, ...], cols: list[int]) -> bool:
    return any(val(row, col) is not None for col in cols)


def any_value(row: tuple[Any, ...], cols: list[int], values: set[int]) -> bool | None:
    if not has_any_response(row, cols):
        return None
    return any(val(row, col) in values for col in cols)


def index_pair(rows: list[tuple[Any, ...]]) -> dict[tuple[str, str], tuple[Any, ...]]:
    return {(clean(row[0]), clean(row[1])): row for row in rows if clean(row[0]) and clean(row[1])}


def condition_key(row: tuple[Any, ...], dimension: str, spec: dict[str, Any], c_meta: dict[int, dict[str, Any]]) -> tuple[str, str] | None:
    value = val(row, spec["column"])
    if value is None:
        return None
    labels = c_meta[spec["column"]]["value_labels"]
    label = labels.get(value, f"code_{value}")
    return f"{dimension}:{value}", f"{spec['label']}={label}"


def support_profile(pairs: list[tuple[tuple[Any, ...], tuple[Any, ...]]], group: str) -> dict[str, Any]:
    b_cols = B_GROUP_COLUMNS[group]
    c_cols = C_GROUP_COLUMNS[group]
    counts: Counter[str] = Counter()
    for b_row, c_row in pairs:
        b_support = any_value(b_row, b_cols, {1})
        b_unmet = any_value(b_row, b_cols, {2})
        c_need = any_value(c_row, c_cols, {1, 2})
        c_strong = any_value(c_row, c_cols, {1})
        if b_support is not None:
            counts["b_usable"] += 1
            counts["b_support_present"] += int(b_support)
        if b_unmet is not None:
            counts["b_unmet_usable"] += 1
            counts["b_needed_not_present"] += int(b_unmet)
        if c_need is not None:
            counts["c_usable"] += 1
            counts["c_need_or_usefulness"] += int(c_need)
        if c_strong is not None:
            counts["c_strong_usable"] += 1
            counts["c_strong_need_or_usefulness"] += int(c_strong)
        if b_support is not None and c_need is not None:
            counts["paired_usable"] += 1
            counts["worker_need_without_b_support"] += int(c_need and not b_support)
            counts["b_support_worker_no_need"] += int(b_support and not c_need)
        if b_unmet is not None and c_need is not None:
            counts["b_unmet_c_need_usable"] += 1
            counts["b_unmet_and_worker_need"] += int(b_unmet and c_need)

    return {
        "group": group,
        "label": SUPPORT_GROUP_LABELS[group],
        "routes": SUPPORT_GROUP_ROUTES[group],
        "freedom_axis": SUPPORT_GROUP_FREEDOM[group],
        "b_usable": counts["b_usable"],
        "c_usable": counts["c_usable"],
        "paired_usable": counts["paired_usable"],
        "b_support_present_rate": rate(counts["b_support_present"], counts["b_usable"]),
        "b_needed_not_present_rate": rate(counts["b_needed_not_present"], counts["b_unmet_usable"]),
        "c_need_or_usefulness_rate": rate(counts["c_need_or_usefulness"], counts["c_usable"]),
        "c_strong_need_or_usefulness_rate": rate(counts["c_strong_need_or_usefulness"], counts["c_strong_usable"]),
        "worker_need_without_b_support_rate": rate(counts["worker_need_without_b_support"], counts["paired_usable"]),
        "b_support_worker_no_need_rate": rate(counts["b_support_worker_no_need"], counts["paired_usable"]),
        "b_unmet_and_worker_need_rate": rate(counts["b_unmet_and_worker_need"], counts["b_unmet_c_need_usable"]),
    }


def contact_profile(pairs: list[tuple[tuple[Any, ...], tuple[Any, ...]]], block: str) -> dict[str, Any]:
    spec = CONTACT_BLOCKS[block]
    counts: Counter[str] = Counter()
    for b_row, _c_row in pairs:
        if not has_any_response(b_row, spec["columns"]):
            continue
        counts["usable"] += 1
        counts["any_resolved"] += int(any_value(b_row, spec["columns"], {1}) is True)
        counts["any_unresolved"] += int(any_value(b_row, spec["columns"], {2}) is True)
        counts["all_no_problem"] += int(all((val(b_row, col) == 3) for col in spec["columns"] if val(b_row, col) is not None))
    return {
        "block": block,
        "label": spec["label"],
        "routes": spec["routes"],
        "freedom_axis": spec["freedom_axis"],
        "usable": counts["usable"],
        "any_resolved_rate": rate(counts["any_resolved"], counts["usable"]),
        "any_unresolved_rate": rate(counts["any_unresolved"], counts["usable"]),
        "all_no_problem_rate": rate(counts["all_no_problem"], counts["usable"]),
    }


def constraint_profile(pairs: list[tuple[tuple[Any, ...], tuple[Any, ...]]], category: str) -> dict[str, Any]:
    spec = CONSTRAINTS[category]
    counts: Counter[str] = Counter()
    for b_row, _c_row in pairs:
        value = val(b_row, spec["column"])
        if value is None:
            continue
        counts["usable"] += 1
        counts["needed_any_constraint"] += int(value in {2, 3})
        counts["needed_serious_constraint"] += int(value == 3)
        counts["not_needed_for_work"] += int(value == 4)
    return {
        "category": category,
        "routes": spec["routes"],
        "freedom_axis": spec["freedom_axis"],
        "usable": counts["usable"],
        "needed_any_constraint_rate": rate(counts["needed_any_constraint"], counts["usable"]),
        "needed_serious_constraint_rate": rate(counts["needed_serious_constraint"], counts["usable"]),
        "not_needed_for_work_rate": rate(counts["not_needed_for_work"], counts["usable"]),
    }


def participation_profile(pairs: list[tuple[tuple[Any, ...], tuple[Any, ...]]]) -> dict[str, Any]:
    counts: Counter[str] = Counter()
    work_cols = {40: "people_work_target", 41: "object_machine_work_target", 42: "data_number_work_target"}
    work_counts: dict[str, Counter[str]] = {name: Counter() for name in work_cols.values()}
    for _b_row, c_row in pairs:
        satisfaction = val(c_row, 105)
        if satisfaction is not None:
            counts["satisfaction_usable"] += 1
            counts["satisfied_or_very_satisfied"] += int(satisfaction in {1, 2})
            counts["dissatisfied_or_not_satisfied"] += int(satisfaction in {4, 5})
        for col, name in work_cols.items():
            value = val(c_row, col)
            if value is None:
                continue
            work_counts[name]["usable"] += 1
            work_counts[name]["yes"] += int(value == 1)
            work_counts[name]["ambiguous"] += int(value == 2)
            work_counts[name]["no"] += int(value == 3)
    return {
        "satisfaction_usable": counts["satisfaction_usable"],
        "satisfied_or_very_satisfied_rate": rate(counts["satisfied_or_very_satisfied"], counts["satisfaction_usable"]),
        "dissatisfied_or_not_satisfied_rate": rate(counts["dissatisfied_or_not_satisfied"], counts["satisfaction_usable"]),
        "work_target_rates": {
            name: {
                "usable": c["usable"],
                "yes_rate": rate(c["yes"], c["usable"]),
                "ambiguous_rate": rate(c["ambiguous"], c["usable"]),
                "no_rate": rate(c["no"], c["usable"]),
            }
            for name, c in work_counts.items()
        },
    }


def mismatch_profile(pairs: list[tuple[tuple[Any, ...], tuple[Any, ...]]]) -> dict[str, Any]:
    counts: Counter[str] = Counter()
    for b_row, c_row in pairs:
        b_disability = val(b_row, 8)
        c_disability = val(c_row, 16)
        if b_disability is not None and c_disability is not None:
            counts["primary_disability_usable"] += 1
            counts["primary_disability_match"] += int(b_disability == c_disability)
        b_severity = val(b_row, 12)
        c_severity = val(c_row, 20)
        if b_severity is not None and c_severity is not None:
            counts["severity_usable"] += 1
            counts["severity_match"] += int(b_severity == c_severity)
    return {
        "primary_disability_match_rate": rate(counts["primary_disability_match"], counts["primary_disability_usable"]),
        "primary_disability_usable": counts["primary_disability_usable"],
        "severity_match_rate": rate(counts["severity_match"], counts["severity_usable"]),
        "severity_usable": counts["severity_usable"],
        "interpretation_boundary": "B/C mismatch is a QA and perspective-alignment signal, not an error assignment.",
    }


def add_deltas(profile: dict[str, Any], baseline: dict[str, Any], rate_keys: list[str]) -> dict[str, Any]:
    out = dict(profile)
    for key in rate_keys:
        out[f"{key}_delta"] = delta(profile.get(key), baseline.get(key))
    return out


def score_support(item: dict[str, Any]) -> float:
    pieces = [
        abs(item.get("c_need_or_usefulness_rate_delta") or 0),
        abs(item.get("b_support_present_rate_delta") or 0),
        abs(item.get("worker_need_without_b_support_rate_delta") or 0),
        abs(item.get("b_needed_not_present_rate_delta") or 0),
        (item.get("worker_need_without_b_support_rate") or 0) * 0.35,
    ]
    return round(sum(pieces), 4)


def score_contact(item: dict[str, Any]) -> float:
    return round(abs(item.get("any_unresolved_rate_delta") or 0) + (item.get("any_unresolved_rate") or 0) * 0.25, 4)


def score_constraint(item: dict[str, Any]) -> float:
    return round(abs(item.get("needed_any_constraint_rate_delta") or 0) + abs(item.get("needed_serious_constraint_rate_delta") or 0), 4)


def summarize_condition(
    window_id: str,
    label: str,
    dimension: str,
    pairs: list[tuple[tuple[Any, ...], tuple[Any, ...]]],
    baselines: dict[str, Any],
) -> dict[str, Any]:
    support_items = [
        add_deltas(
            support_profile(pairs, group),
            baselines["support"][group],
            [
                "b_support_present_rate",
                "b_needed_not_present_rate",
                "c_need_or_usefulness_rate",
                "c_strong_need_or_usefulness_rate",
                "worker_need_without_b_support_rate",
                "b_support_worker_no_need_rate",
                "b_unmet_and_worker_need_rate",
            ],
        )
        for group in B_GROUP_COLUMNS
    ]
    for item in support_items:
        item["signal_score"] = score_support(item)

    contact_items = [
        add_deltas(contact_profile(pairs, block), baselines["contact"][block], ["any_resolved_rate", "any_unresolved_rate", "all_no_problem_rate"])
        for block in CONTACT_BLOCKS
    ]
    for item in contact_items:
        item["signal_score"] = score_contact(item)

    constraint_items = [
        add_deltas(
            constraint_profile(pairs, category),
            baselines["constraint"][category],
            ["needed_any_constraint_rate", "needed_serious_constraint_rate", "not_needed_for_work_rate"],
        )
        for category in CONSTRAINTS
    ]
    for item in constraint_items:
        item["signal_score"] = score_constraint(item)

    participation = participation_profile(pairs)
    participation["satisfied_or_very_satisfied_rate_delta"] = delta(
        participation.get("satisfied_or_very_satisfied_rate"),
        baselines["participation"].get("satisfied_or_very_satisfied_rate"),
    )
    participation["dissatisfied_or_not_satisfied_rate_delta"] = delta(
        participation.get("dissatisfied_or_not_satisfied_rate"),
        baselines["participation"].get("dissatisfied_or_not_satisfied_rate"),
    )
    for name, rates in participation["work_target_rates"].items():
        base = baselines["participation"]["work_target_rates"][name]
        rates["yes_rate_delta"] = delta(rates.get("yes_rate"), base.get("yes_rate"))
        rates["ambiguous_rate_delta"] = delta(rates.get("ambiguous_rate"), base.get("ambiguous_rate"))

    top_support = sorted(support_items, key=lambda item: item["signal_score"], reverse=True)[:5]
    top_contact = sorted(contact_items, key=lambda item: item["signal_score"], reverse=True)[:3]
    top_constraint = sorted(constraint_items, key=lambda item: item["signal_score"], reverse=True)[:5]

    routes: Counter[str] = Counter()
    axes: Counter[str] = Counter()
    for item in top_support[:3]:
        routes.update(item["routes"])
        axes.update([item["freedom_axis"]])
    for item in top_contact[:2]:
        routes.update(item["routes"])
        axes.update([item["freedom_axis"]])
    for item in top_constraint[:3]:
        routes.update(item["routes"])
        axes.update([item["freedom_axis"]])

    return {
        "window_id": window_id,
        "dimension": dimension,
        "label": label,
        "linked_pair_count": len(pairs),
        "status": "condition_window_interaction_profile_unreviewed_no_promotion",
        "mismatch_QA": mismatch_profile(pairs),
        "top_support_translation_signals": top_support,
        "top_contact_problem_signals": top_contact,
        "top_occupational_constraint_signals": top_constraint,
        "participation_signals": participation,
        "route_counts": dict(sorted(routes.items())),
        "freedom_axis_counts": dict(sorted(axes.items())),
        "candidate_interaction_proposition": candidate_interaction_proposition(label, top_support, top_contact, top_constraint, participation),
        "counter_proposition": "The observed association may reflect survey design, job assignment, workplace selection, B/C alignment differences, era-specific employment composition, or missing unobserved context rather than a condition-specific mechanism.",
        "human_review_question": "Which function, activity, contact point, support translation, time/load, disclosure, evaluation, or life-security mechanism explains this condition-window signal, and what competing explanation remains?",
        "not_allowed": [
            "disease/disability-to-support lookup",
            "worker capacity decision",
            "support adequacy decision",
            "medical/legal/employment judgment",
            "current-policy claim",
            "candidate_pattern movement",
            "public/runtime approval",
        ],
        "source_content_exported": False,
        "narrative_content_included": False,
        "row_level_ids_exported": False,
    }


def direction(value: float | None) -> str:
    if value is None:
        return "unknown"
    if value >= 0.05:
        return "higher-than-baseline"
    if value <= -0.05:
        return "lower-than-baseline"
    return "near-baseline"


def candidate_interaction_proposition(
    label: str,
    support_items: list[dict[str, Any]],
    contact_items: list[dict[str, Any]],
    constraint_items: list[dict[str, Any]],
    participation: dict[str, Any],
) -> str:
    support = support_items[0]
    contact = contact_items[0]
    constraint = constraint_items[0]
    parts = [
        f"{label}は、それ自体を説明原因にするのではなく、機序探索の条件窓として使う。",
        f"最も強い支援翻訳シグナルは `{support['group']}` で、本人側の必要/有用性は {direction(support.get('c_need_or_usefulness_rate_delta'))}、本人ニーズがあるがB側支援ありになっていないギャップは {direction(support.get('worker_need_without_b_support_rate_delta'))}。",
        f"最も強い職場接触点シグナルは `{contact['block']}` / {contact['label']} で、未解決問題率は {direction(contact.get('any_unresolved_rate_delta'))}。",
        f"最も強い職業生活制約シグナルは `{constraint['category']}` で、何らかの制約率は {direction(constraint.get('needed_any_constraint_rate_delta'))}。",
    ]
    if participation.get("dissatisfied_or_not_satisfied_rate_delta") is not None:
        parts.append(
            f"不満足側の差は {direction(participation.get('dissatisfied_or_not_satisfied_rate_delta'))} なので、参加品質との接続も点検する。"
        )
    return " ".join(parts)


def build_outputs() -> dict[str, Any]:
    c_meta = dictionary("C_data.xlsx")
    b_rows = index_pair(load_rows("B_data.xlsx", "B全て"))
    c_rows = index_pair(load_rows("C_data.xlsx", "C全て"))
    linked_keys = sorted(set(b_rows) & set(c_rows))
    all_pairs = [(b_rows[key], c_rows[key]) for key in linked_keys]

    baselines = {
        "support": {group: support_profile(all_pairs, group) for group in B_GROUP_COLUMNS},
        "contact": {block: contact_profile(all_pairs, block) for block in CONTACT_BLOCKS},
        "constraint": {category: constraint_profile(all_pairs, category) for category in CONSTRAINTS},
        "participation": participation_profile(all_pairs),
        "mismatch_QA": mismatch_profile(all_pairs),
    }

    buckets: dict[tuple[str, str, str], list[tuple[tuple[Any, ...], tuple[Any, ...]]]] = defaultdict(list)
    for key in linked_keys:
        pair = (b_rows[key], c_rows[key])
        c_row = c_rows[key]
        for dimension, spec in CONDITION_WINDOWS.items():
            item = condition_key(c_row, dimension, spec, c_meta)
            if item is None:
                continue
            window_id, label = item
            buckets[(dimension, window_id, label)].append(pair)

    profiles: list[dict[str, Any]] = []
    for (dimension, window_id, label), pairs in sorted(buckets.items()):
        min_n = CONDITION_WINDOWS[dimension]["min_n"]
        if len(pairs) < min_n:
            continue
        profiles.append(summarize_condition(window_id, label, dimension, pairs, baselines))

    profiles.sort(key=lambda item: (item["dimension"], -item["linked_pair_count"], item["window_id"]))

    interaction_cards = build_interaction_cards(profiles)

    return {
        "dataset_id": DATASET_ID,
        "artifact_id": "2001_ABC_survey_condition_window_interaction_deepening_v0_2026_05_23",
        "lane": "Falcon Lab",
        "status": "condition_window_interaction_deepening_unreviewed_no_text_no_promotion",
        "source_layers_used": [
            "data/original_secure/structured/2001_ABC_survey/B_data.xlsx",
            "data/original_secure/structured/2001_ABC_survey/C_data.xlsx",
        ],
        "source_content_exported": False,
        "narrative_content_included": False,
        "row_level_ids_exported": False,
        "linked_B_C_pair_count": len(linked_keys),
        "baseline_profiles": baselines,
        "condition_window_profile_count": len(profiles),
        "condition_window_profiles": profiles,
        "interaction_card_count": len(interaction_cards),
        "interaction_cards": interaction_cards,
        "boundary": [
            "Condition-window relations are allowed as unreviewed interaction evidence.",
            "Do not convert condition labels into support lookup rules.",
            "Do not infer worker capacity, support adequacy, employer accuracy, or current-policy guidance.",
            "Human review is required before candidate_pattern, reviewed knowledge, public, or runtime use.",
        ],
        "review_status": "unreviewed",
    }


def build_interaction_cards(profiles: list[dict[str, Any]]) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for profile in profiles:
        if profile["linked_pair_count"] < 60:
            continue
        for signal in profile["top_support_translation_signals"][:3]:
            if signal["signal_score"] < 0.08:
                continue
            candidates.append(
                {
                    "kind": "support_translation",
                    "score": signal["signal_score"],
                    "profile": profile,
                    "signal": signal,
                    "routes": signal["routes"],
                    "freedom_axes": [signal["freedom_axis"]],
                }
            )
        for signal in profile["top_contact_problem_signals"][:2]:
            if signal["signal_score"] < 0.08:
                continue
            candidates.append(
                {
                    "kind": "workplace_contact",
                    "score": signal["signal_score"],
                    "profile": profile,
                    "signal": signal,
                    "routes": signal["routes"],
                    "freedom_axes": [signal["freedom_axis"]],
                }
            )
        for signal in profile["top_occupational_constraint_signals"][:2]:
            if signal["signal_score"] < 0.08:
                continue
            candidates.append(
                {
                    "kind": "occupational_constraint",
                    "score": signal["signal_score"],
                    "profile": profile,
                    "signal": signal,
                    "routes": signal["routes"],
                    "freedom_axes": [signal["freedom_axis"]],
                }
            )

    selected = sorted(candidates, key=lambda item: item["score"], reverse=True)[:40]
    cards: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()
    for idx, item in enumerate(selected, start=1):
        profile = item["profile"]
        signal = item["signal"]
        signal_name = signal.get("group") or signal.get("block") or signal.get("category")
        key = (profile["window_id"], item["kind"], signal_name)
        if key in seen:
            continue
        seen.add(key)
        card_id = f"2001_ABC_survey:condition-window-interaction:{idx:03d}"
        cards.append(
            {
                "card_id": card_id,
                "dataset_id": DATASET_ID,
                "status": "condition_window_interaction_card_unreviewed_no_promotion",
                "condition_window": {
                    "window_id": profile["window_id"],
                    "dimension": profile["dimension"],
                    "label": profile["label"],
                    "linked_pair_count": profile["linked_pair_count"],
                },
                "signal_kind": item["kind"],
                "signal": signal,
                "routes": item["routes"],
                "freedom_axes": item["freedom_axes"],
                "activation_terms": activation_terms(profile, signal, item["kind"]),
                "candidate_interaction_proposition": card_proposition(profile, signal, item["kind"]),
                "counter_proposition": profile["counter_proposition"],
                "human_review_question": profile["human_review_question"],
                "not_allowed": profile["not_allowed"],
                "source_content_exported": False,
                "narrative_content_included": False,
                "row_level_ids_exported": False,
                "review_status": "unreviewed",
            }
        )
    return cards


def activation_terms(profile: dict[str, Any], signal: dict[str, Any], kind: str) -> list[str]:
    terms = [profile["label"], "条件窓", "相互作用", "SCIMA/FCHMA"]
    if kind == "support_translation":
        terms.extend(["支援", "配慮", SUPPORT_GROUP_LABELS[signal["group"]]])
    elif kind == "workplace_contact":
        terms.extend(["職場接触点", signal["label"]])
    else:
        terms.extend(["職業生活制約", signal["category"]])
    return terms


def card_proposition(profile: dict[str, Any], signal: dict[str, Any], kind: str) -> str:
    label = profile["label"]
    if kind == "support_translation":
        return (
            f"{label}では、`{signal['group']}`を配慮検索キーではなく支援翻訳機序として起動する。"
            f"C側の必要/有用性 {signal.get('c_need_or_usefulness_rate')}、B側の支援あり {signal.get('b_support_present_rate')}、"
            f"本人ニーズがあるがB側支援ありになっていないギャップ {signal.get('worker_need_without_b_support_rate')} を並べて読む。"
        )
    if kind == "workplace_contact":
        return (
            f"{label}では、`{signal['block']}`を職場接触点の機序探索として起動する。"
            f"未解決問題率 {signal.get('any_unresolved_rate')} は接触点シグナルであり、条件窓による運命論ではない。"
        )
    return (
        f"{label}では、`{signal['category']}`を職業生活制約の機序探索として起動する。"
        f"何らかの制約率 {signal.get('needed_any_constraint_rate')} は、仕事要求、支援翻訳、参加文脈と一緒に読む。"
    )


def write_markdown(data: dict[str, Any]) -> None:
    lines = [
        "# 2001 ABC Survey Condition-Window Interaction Deepening",
        "",
        "作成日: 2026-05-23",
        "Lane: Falcon Lab",
        "状態: condition-window interaction deepening / coded structured values only / no narrative text / 未レビュー / 昇格なし / runtime未承認",
        "本文引用: なし",
        "",
        "## Position",
        "",
        "この成果物は、条件窓をholdとして封印せず、病名・障害名・程度・機能障害・回答条件を、支援、困難、職場接触点、参加品質との相互作用を読む条件変数として使うための中間レイヤーである。",
        "",
        "禁止しているのは、条件名から配慮や就労困難性を直接引くこと、支援妥当性を決めること、能力判断をすること、現在の政策・実務指針にすること、レビュー前に昇格することである。",
        "",
        "## Scope",
        "",
        f"- B/C linked coded pairs: {data['linked_B_C_pair_count']}",
        f"- condition-window profiles: {data['condition_window_profile_count']}",
        f"- interaction cards: {data['interaction_card_count']}",
        "",
        "## Top Interaction Cards",
        "",
        "| card | condition window | signal | routes | proposition |",
        "|---|---|---|---|---|",
    ]
    for card in data["interaction_cards"][:24]:
        signal = card["signal"].get("group") or card["signal"].get("block") or card["signal"].get("category")
        routes = ", ".join(f"`{route}`" for route in card["routes"])
        lines.append(
            f"| `{card['card_id']}` | {card['condition_window']['label']} | `{card['signal_kind']}:{signal}` | {routes} | {card['candidate_interaction_proposition']} |"
        )

    lines.extend(
        [
            "",
            "## Condition-Window Profiles",
            "",
            "| window | pairs | top support signal | top contact signal | top constraint signal | review question |",
            "|---|---:|---|---|---|---|",
        ]
    )
    for profile in sorted(data["condition_window_profiles"], key=lambda item: item["linked_pair_count"], reverse=True)[:36]:
        support = profile["top_support_translation_signals"][0]
        contact = profile["top_contact_problem_signals"][0]
        constraint = profile["top_occupational_constraint_signals"][0]
        lines.append(
            f"| {profile['label']} | {profile['linked_pair_count']} | `{support['group']}` score {support['signal_score']} | `{contact['block']}` score {contact['signal_score']} | `{constraint['category']}` score {constraint['signal_score']} | {profile['human_review_question']} |"
        )

    lines.extend(
        [
            "",
            "## Reading Boundary",
            "",
            "- 条件窓関係は、未レビューの相互作用証拠として分析してよい。",
            "- 条件名を配慮検索キー、就労困難性の決定因、能力判断、支援妥当性判断にはしない。",
            "- 観測された差は、調査票設計、職務配置、職場選択、B/C認識差、2001年時点の制度・雇用構成、未観測文脈で説明される可能性を残す。",
            "- `candidate_pattern`、reviewed knowledge、public-safe、public-approved、runtime-approved への移動はしない。",
            "",
        ]
    )
    OUTPUT_MD.write_text("\n".join(lines), encoding="utf-8")


def write_chat_cards(data: dict[str, Any]) -> None:
    with CHAT_CARDS_JSONL.open("w", encoding="utf-8") as out:
        for card in data["interaction_cards"]:
            out.write(json.dumps(card, ensure_ascii=False) + "\n")


def main() -> None:
    data = build_outputs()
    OUTPUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_markdown(data)
    write_chat_cards(data)
    print(
        json.dumps(
            {
                "linked_B_C_pair_count": data["linked_B_C_pair_count"],
                "condition_window_profile_count": data["condition_window_profile_count"],
                "interaction_card_count": data["interaction_card_count"],
                "output": str(OUTPUT_JSON.relative_to(REPO_ROOT)),
                "chat_cards": str(CHAT_CARDS_JSONL.relative_to(REPO_ROOT)),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
