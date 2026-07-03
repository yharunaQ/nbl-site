#!/usr/bin/env python3
"""Read raw original workbooks locally and export no-text context signals.

This script intentionally opens restricted raw original workbooks, but it never
prints or writes raw cell values. It scans only the free-text columns declared in
source manifests and the RR-01..RR-05 workset records.
"""

from __future__ import annotations

from collections import Counter, defaultdict
import json
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = ROOT / "references/derived/scima-fchma/stage1-production-v0-2026-05-18"
PREFIX = "stage1-production-raw-original-local-only-audit-v0-2026-05-23"

WORKSET = OUT_DIR / "stage1-production-raw-redacted-rereading-workset-manifest-v0-2026-05-23.json"
CR_FILES = {
    "RR-01-health-time-life-security": OUT_DIR / "stage1-production-cr01-health-time-life-security-context-reading-v0-2026-05-23.json",
    "RR-02-quality-participation-value-translation": OUT_DIR / "stage1-production-cr02-quality-value-context-reading-v0-2026-05-23.json",
    "RR-03-prework-entry-sequence": OUT_DIR / "stage1-production-cr03-prework-entry-context-reading-v0-2026-05-23.json",
    "RR-04-worksite-contact-minority-window": OUT_DIR / "stage1-production-cr04-worksite-contact-context-reading-v0-2026-05-23.json",
    "RR-05-residual-hold-and-counterexamples": OUT_DIR / "stage1-production-cr05-residual-hold-context-reading-v0-2026-05-23.json",
}

SOURCE_MANIFESTS = {
    "employment_survey_3000": ROOT / "data/specs/source-manifests/employment_survey_3000.source-manifest.json",
    "nanbyo_survey_4000": ROOT / "data/specs/source-manifests/nanbyo_survey_4000.source-manifest.json",
}

RAW_WORKBOOKS = {
    "employment_survey_3000": {
        "path": ROOT / "data/original_secure/structured/employment_survey_3000/disability_employment.xlsx",
        "sheet": "回答データ",
    },
    "nanbyo_survey_4000": {
        "path": ROOT / "data/original_secure/structured/nanbyo_survey_4000/Nanbyo_kanja.xlsx",
        "sheet": "難病患者調査標準データ",
    },
}

LEXICONS = {
    "health_time": ["体調", "疲労", "疲れ", "通院", "治療", "休", "睡眠", "症状", "病状", "悪化", "回復", "体力", "健康"],
    "work_time_design": ["勤務", "時間", "短時間", "休憩", "休暇", "欠勤", "遅刻", "早退", "シフト", "残業"],
    "task_load_design": ["仕事量", "作業", "業務", "職務", "負担", "責任", "ペース", "配置", "仕事内容"],
    "life_security": ["収入", "生活", "給与", "賃金", "年金", "手当", "制度", "保障", "経済"],
    "support_bridge": ["支援", "相談", "上司", "同僚", "職場", "医師", "ハローワーク", "福祉", "ジョブコーチ"],
    "sequence_or_choice": ["継続", "退職", "復職", "転職", "再就職", "将来", "選択", "希望"],
    "role_or_work_value": ["役割", "責任", "成果", "技能", "能力", "仕事", "業務", "職務"],
    "evaluation_or_treatment": ["評価", "処遇", "賃金", "給与", "昇進", "キャリア", "待遇"],
    "future_outlook": ["将来", "継続", "定着", "転職", "退職", "見通し", "希望"],
    "satisfaction_only_risk": ["満足", "やりがい", "不満"],
    "conditional_performance": ["配慮", "支援", "体調", "疲労", "障害", "休", "仕事量"],
    "entry_action": ["応募", "採用", "求人", "面接", "就職活動", "求職", "就職"],
    "prework_training": ["訓練", "実習", "職場見学", "体験", "講座", "練習", "資格", "学習"],
    "life_rhythm_stamina": ["生活リズム", "日中", "体力", "疲", "通院", "治療", "健康", "睡眠"],
    "self_outlook": ["自信", "希望", "不安", "将来", "できる", "能力"],
    "nonwork_or_low_context_brake": ["わからない", "不明", "働きたく", "希望しない", "無理"],
    "body_or_function_boundary": ["身体", "機能", "障害", "痛", "しびれ", "視覚", "聴覚", "精神", "知的"],
}

CORRIDOR_REQUIRED_TYPES = {
    "RR-01-health-time-life-security": [
        {"health_time", "life_security"},
        {"health_time", "work_time_design"},
        {"health_time", "task_load_design"},
        {"life_security", "support_bridge", "sequence_or_choice"},
        {"health_time", "support_bridge", "sequence_or_choice"},
    ],
    "RR-02-quality-participation-value-translation": [
        {"role_or_work_value", "conditional_performance"},
        {"evaluation_or_treatment", "role_or_work_value", "support_bridge"},
        {"future_outlook", "role_or_work_value"},
        {"future_outlook", "evaluation_or_treatment"},
    ],
    "RR-03-prework-entry-sequence": [
        {"entry_action", "prework_training"},
        {"life_rhythm_stamina", "support_bridge"},
        {"self_outlook", "entry_action"},
        {"self_outlook", "prework_training"},
    ],
    "RR-04-worksite-contact-minority-window": [
        {"task_load_design", "support_bridge"},
        {"work_time_design", "support_bridge"},
        {"body_or_function_boundary", "task_load_design"},
    ],
    "RR-05-residual-hold-and-counterexamples": [],
}


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def normalize_id(value: Any) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    if text.endswith(".0"):
        text = text[:-2]
    if text.isdigit():
        return str(int(text))
    return None


def split_record_id(record_id: str) -> tuple[str, str]:
    dataset_id, local = record_id.split(":", 1)
    return dataset_id, str(int(local))


def load_workset_records() -> list[dict[str, str]]:
    workset = load_json(WORKSET)
    rows: list[dict[str, str]] = []
    for ws in workset["worksets"]:
        corridor = ws["corridor"]
        for group, ids in ws["record_groups"].items():
            for rid in ids:
                rows.append({"record_id": rid, "corridor": corridor, "group": group})
    return rows


def load_free_text_columns() -> dict[str, list[str]]:
    out = {}
    for dataset_id, path in SOURCE_MANIFESTS.items():
        manifest = load_json(path)
        out[dataset_id] = list(manifest["free_text_columns"])
    return out


def detect_types(text: str) -> dict[str, int]:
    out = {}
    for ctype, terms in LEXICONS.items():
        hits = sum(1 for term in terms if term in text)
        if hits:
            out[ctype] = hits
    return out


def char_bucket(total_chars: int) -> str:
    if total_chars == 0:
        return "0"
    if total_chars < 80:
        return "1-79"
    if total_chars < 250:
        return "80-249"
    if total_chars < 800:
        return "250-799"
    return "800+"


def load_raw_contexts(workset_rows: list[dict[str, str]], free_text_columns: dict[str, list[str]]) -> dict[str, dict[str, Any]]:
    targets: dict[str, set[str]] = defaultdict(set)
    for row in workset_rows:
        dataset_id, local = split_record_id(row["record_id"])
        targets[dataset_id].add(local)

    contexts: dict[str, dict[str, Any]] = {}
    for dataset_id, ids in targets.items():
        spec = RAW_WORKBOOKS[dataset_id]
        workbook = openpyxl.load_workbook(spec["path"], read_only=True, data_only=True)
        sheet = workbook[spec["sheet"]]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        header_index = {str(name): idx for idx, name in enumerate(headers) if name is not None}
        id_index = header_index["ID"]
        scan_columns = [col for col in free_text_columns[dataset_id] if col in header_index]
        missing_columns = sorted(set(free_text_columns[dataset_id]) - set(scan_columns))
        scan_indexes = [(col, header_index[col]) for col in scan_columns]

        for cells in sheet.iter_rows(min_row=2, values_only=True):
            local = normalize_id(cells[id_index])
            if local not in ids:
                continue
            rid = f"{dataset_id}:{int(local):05d}"
            present_fields = 0
            total_chars = 0
            combined_type_counts = Counter()
            for _col, idx in scan_indexes:
                value = cells[idx]
                if value is None:
                    continue
                text = str(value).strip()
                if not text:
                    continue
                present_fields += 1
                total_chars += len(text)
                combined_type_counts.update(detect_types(text))
            contexts[rid] = {
                "dataset_id": dataset_id,
                "raw_row_found": True,
                "free_text_fields_declared": len(free_text_columns[dataset_id]),
                "free_text_fields_scanned": len(scan_columns),
                "free_text_fields_missing_from_workbook": missing_columns,
                "raw_free_text_fields_present": present_fields,
                "raw_free_text_char_bucket": char_bucket(total_chars),
                "raw_context_type_counts_without_quotes": dict(sorted(combined_type_counts.items())),
                "raw_text_exported": False,
            }
        workbook.close()
    return contexts


def load_prior_cr_actions() -> dict[str, dict[str, str]]:
    prior: dict[str, dict[str, str]] = {}
    for path in CR_FILES.values():
        data = load_json(path)
        for row in data["record_context_notes"]:
            prior[row["record_id"]] = {
                "prior_action": row["context_reading_action"],
                "prior_redacted_unit_available": row.get("redacted_unit_available", True),
            }
    return prior


def closes_required_relation(corridor: str, context_types: set[str]) -> bool:
    return any(required <= context_types for required in CORRIDOR_REQUIRED_TYPES.get(corridor, []))


def build() -> dict[str, Any]:
    workset_rows = load_workset_records()
    free_text_columns = load_free_text_columns()
    raw_contexts = load_raw_contexts(workset_rows, free_text_columns)
    prior = load_prior_cr_actions()

    records = []
    corridor_counts: dict[str, Counter[str]] = defaultdict(Counter)
    raw_type_counts: dict[str, Counter[str]] = defaultdict(Counter)
    raw_adds_closure = Counter()
    missing_raw_rows = []
    for row in workset_rows:
        rid = row["record_id"]
        raw = raw_contexts.get(rid)
        if raw is None:
            raw = {
                "dataset_id": rid.split(":", 1)[0],
                "raw_row_found": False,
                "free_text_fields_declared": 0,
                "free_text_fields_scanned": 0,
                "free_text_fields_missing_from_workbook": [],
                "raw_free_text_fields_present": 0,
                "raw_free_text_char_bucket": "0",
                "raw_context_type_counts_without_quotes": {},
                "raw_text_exported": False,
            }
            missing_raw_rows.append(rid)
        context_types = set(raw["raw_context_type_counts_without_quotes"])
        closure = closes_required_relation(row["corridor"], context_types)
        prior_action = prior.get(rid, {}).get("prior_action", "not_in_prior_cr")
        if closure and prior_action.startswith("structured_coverage"):
            raw_adds_closure[row["corridor"]] += 1
        corridor_counts[row["corridor"]][
            "records"
        ] += 1
        corridor_counts[row["corridor"]][
            "raw_rows_found" if raw["raw_row_found"] else "raw_rows_missing"
        ] += 1
        corridor_counts[row["corridor"]][
            "raw_free_text_present" if raw["raw_free_text_fields_present"] else "no_raw_free_text_present"
        ] += 1
        corridor_counts[row["corridor"]][
            "raw_closes_required_relation" if closure else "raw_does_not_close_required_relation"
        ] += 1
        raw_type_counts[row["corridor"]].update(context_types)
        records.append(
            {
                "record_id": rid,
                "corridor": row["corridor"],
                "group": row["group"],
                "reading_surface": "raw_original_local_only",
                "raw_original_opened": True,
                "source_text_exported": False,
                "redacted_text_exported": False,
                "field_value_exported": False,
                "llm_received_raw_text": False,
                **raw,
                "raw_closes_required_relation": closure,
                "prior_context_reading_action": prior_action,
                "raw_effect_on_prior_action": (
                    "raw_adds_possible_closure"
                    if closure and prior_action.startswith("structured_coverage")
                    else "raw_confirms_or_keeps_prior_boundary"
                ),
            }
        )

    payload = {
        "artifact_id": PREFIX,
        "lane": "Falcon / Falcon Lab",
        "status": "raw_original_local_only_audit_no_text_export_no_promotion",
        "review_status": "unreviewed",
        "promotion_status": "none",
        "public_status": "not_public",
        "runtime_status": "not_runtime_approved",
        "raw_original_opened": True,
        "source_text_exported": False,
        "redacted_text_exported": False,
        "field_value_exported": False,
        "llm_received_raw_text": False,
        "target_scope": {
            "corpus": sorted(RAW_WORKBOOKS),
            "records": "RR-01..RR-05 workset records only",
            "fields": free_text_columns,
            "purpose": "SCIMA/FCHMA context closure check for Stage 1 rereading gaps, especially C07/C08",
            "stop_condition": "emit no-text route effect summary; do not inspect beyond declared free-text columns",
        },
        "record_count": len(records),
        "unique_record_count": len({row["record_id"] for row in records}),
        "record_corridor_entry_count": len(records),
        "missing_raw_rows": missing_raw_rows,
        "corridor_counts": {key: dict(value) for key, value in sorted(corridor_counts.items())},
        "raw_context_type_record_counts_by_corridor": {
            key: dict(sorted(value.items())) for key, value in sorted(raw_type_counts.items())
        },
        "raw_adds_possible_closure_from_structured_coverage": dict(sorted(raw_adds_closure.items())),
        "records": records,
    }
    return payload


def make_md(data: dict[str, Any]) -> str:
    lines = [
        "# Stage 1 Raw Original Local-Only Audit",
        "",
        "作成日: 2026-05-23",
        "Lane: Falcon / Falcon Lab",
        "状態: raw original local-only audit / 本文引用なし / 昇格なし / 公開不可",
        "",
        "RR-01..RR-05 worksetの対象recordについて、raw original workbookをlocal-onlyで開いた。",
        "出力にはraw本文、伏字段落、field値を含めない。LLMにもraw本文を渡していない。",
        "",
        "## Scope",
        f"- raw_original_opened: {data['raw_original_opened']}",
        f"- source_text_exported: {data['source_text_exported']}",
        f"- llm_received_raw_text: {data['llm_received_raw_text']}",
        f"- unique records: {data['unique_record_count']}",
        f"- record-corridor entries: {data['record_corridor_entry_count']}",
        "",
        "## Corridor Counts",
        "| corridor | records | raw rows found | raw free-text present | raw closes required relation |",
        "|---|---:|---:|---:|---:|",
    ]
    for corridor, counts in data["corridor_counts"].items():
        lines.append(
            f"| {corridor} | {counts.get('records', 0)} | {counts.get('raw_rows_found', 0)} | "
            f"{counts.get('raw_free_text_present', 0)} | {counts.get('raw_closes_required_relation', 0)} |"
        )
    lines.extend([
        "",
        "## Raw Adds Possible Closure From Structured Coverage",
        "",
        "structured coverage止まりだったrecordのうち、raw local-only文脈型で必要関係が閉じる可能性が出た件数。",
        "",
        "| corridor | records |",
        "|---|---:|",
    ])
    for corridor, count in data["raw_adds_possible_closure_from_structured_coverage"].items():
        lines.append(f"| {corridor} | {count} |")
    if not data["raw_adds_possible_closure_from_structured_coverage"]:
        lines.append("| none | 0 |")
    lines.extend([
        "",
        "## Interpretation",
        "- raw originalは開いたが、本文は外に出していない。",
        "- C07/C08を厚くする場合は、このauditで`raw_adds_possible_closure`になったrecordだけを次のno-text再分類に送る。",
        "- source/support/intervention validity、医療・雇用・配慮妥当性、review statusは動かさない。",
        "",
    ])
    return "\n".join(lines)


def main() -> None:
    data = build()
    (OUT_DIR / f"{PREFIX}.json").write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (OUT_DIR / f"{PREFIX}-records.jsonl").write_text(
        "".join(json.dumps(row, ensure_ascii=False) + "\n" for row in data["records"]),
        encoding="utf-8",
    )
    (OUT_DIR / f"{PREFIX}.md").write_text(make_md(data), encoding="utf-8")
    print(PREFIX, "records=", data["record_count"], "raw_original_opened=", data["raw_original_opened"])


if __name__ == "__main__":
    main()
