#!/usr/bin/env python3
"""Build no-text A and C single-perspective SCIMA/FCHMA profiles.

This script keeps A and C valuable in their own right instead of treating them
only as linkage supports. It reads structured coded values from original_secure
and writes aggregate, no-row-text profiles under references/derived.
"""

from __future__ import annotations

import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
DATASET_ID = "2001_ABC_survey"
SOURCE_DIR = REPO_ROOT / "data/original_secure/structured/2001_ABC_survey"
DERIVED_DIR = REPO_ROOT / "references/derived/scima-fchma/2001-abc-survey-v0-2026-05-22"

OUTPUT_JSON = DERIVED_DIR / "2001-abc-survey-A-C-single-perspective-structure-profiles-v0-2026-05-22.json"
OUTPUT_MD = DERIVED_DIR / "2001-abc-survey-A-C-single-perspective-structure-profiles-v0-2026-05-22.md"


A_BURDEN_COLUMNS = list(range(41, 55))
A_ADVICE_COLUMNS = list(range(65, 92, 2))
A_EMPLOYMENT_REASON_COLUMNS = [55, 56, 57, 58]
A_EMPLOYMENT_CHALLENGE_COLUMNS = [60, 61, 62, 63]

C_MAIN_CONDITION_COLUMNS = [16, 17, 18, 20, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 38]
C_WORK_CONTENT_COLUMNS = [40, 41, 42]
C_SUPPORT_COLUMNS = list(range(44, 102))
C_SATISFACTION_COLUMN = 105
C_RESPONSE_METHOD_COLUMNS = [4, 5, 6]


SUPPORT_GROUPS = {
    "training_skill_education": list(range(44, 48)),
    "job_task_design_and_performance": list(range(48, 54)),
    "stress_responsibility_safety": list(range(54, 59)),
    "information_communication": list(range(59, 64)),
    "within_workplace_mobility": list(range(64, 71)),
    "commuting": list(range(71, 80)),
    "health_time_self_care_environment": list(range(80, 88)),
    "interpersonal_relations_and_workplace_culture": list(range(88, 94)),
    "emergency_disaster": list(range(94, 97)),
    "offwork_daily_living": list(range(97, 102)),
}


A_ITEM_ROUTES = {
    "建築物の改造": ["QR-03-worksite-contact-and-mobility"],
    "作業設備の改善": ["QR-03-worksite-contact-and-mobility"],
    "支援機器の整備": ["QR-03-worksite-contact-and-mobility", "QR-02-information-work-procedure"],
    "専任の職場支援・介護者の配置": ["QR-06-disclosure-boundary-and-mutual-translation", "QR-07-quality-career-and-value-translation"],
    "他の従業員による支援・介助の促進": ["QR-06-disclosure-boundary-and-mutual-translation", "QR-07-quality-career-and-value-translation"],
    "管理職への障害者雇用啓発": ["QR-06-disclosure-boundary-and-mutual-translation"],
    "従業員への障害者雇用啓発": ["QR-06-disclosure-boundary-and-mutual-translation", "QR-07-quality-career-and-value-translation"],
    "障害者の特性に応じた職場配置": ["QR-03-worksite-contact-and-mobility", "QR-07-quality-career-and-value-translation"],
    "教育・訓練": ["QR-05-entry-prework-translation", "QR-07-quality-career-and-value-translation"],
    "安全面の配慮や環境改善": ["QR-03-worksite-contact-and-mobility"],
    "健康管理": ["QR-01-health-time-work-design"],
    "給与・労働条件の改善": ["QR-04-life-security-sequencing", "QR-07-quality-career-and-value-translation"],
    "障害者の勤務時間外の自立生活支援": ["QR-04-life-security-sequencing"],
    "家族や学校・福祉施設との連携": ["QR-04-life-security-sequencing", "QR-06-disclosure-boundary-and-mutual-translation"],
}


SUPPORT_GROUP_ROUTES = {
    "training_skill_education": ["QR-05-entry-prework-translation", "QR-07-quality-career-and-value-translation"],
    "job_task_design_and_performance": ["QR-03-worksite-contact-and-mobility", "QR-07-quality-career-and-value-translation"],
    "stress_responsibility_safety": ["QR-01-health-time-work-design", "QR-03-worksite-contact-and-mobility", "QR-07-quality-career-and-value-translation"],
    "information_communication": ["QR-02-information-work-procedure", "QR-06-disclosure-boundary-and-mutual-translation"],
    "within_workplace_mobility": ["QR-03-worksite-contact-and-mobility"],
    "commuting": ["QR-03-worksite-contact-and-mobility", "QR-04-life-security-sequencing"],
    "health_time_self_care_environment": ["QR-01-health-time-work-design", "QR-04-life-security-sequencing"],
    "interpersonal_relations_and_workplace_culture": ["QR-06-disclosure-boundary-and-mutual-translation", "QR-07-quality-career-and-value-translation"],
    "emergency_disaster": ["QR-03-worksite-contact-and-mobility", "QR-08-diversity-conditioned-same-structure"],
    "offwork_daily_living": ["QR-04-life-security-sequencing", "QR-06-disclosure-boundary-and-mutual-translation"],
}


SUPPORT_GROUP_FREEDOM = {
    "training_skill_education": "learning_variability_freedom",
    "job_task_design_and_performance": "work_content_freedom",
    "stress_responsibility_safety": "load_quality_safety_freedom",
    "information_communication": "information_translation_freedom",
    "within_workplace_mobility": "body_environment_freedom",
    "commuting": "life_commuting_freedom",
    "health_time_self_care_environment": "health_time_freedom",
    "interpersonal_relations_and_workplace_culture": "social_participation_freedom",
    "emergency_disaster": "safety_contingency_freedom",
    "offwork_daily_living": "life_security_freedom",
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def numeric(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    try:
        number = int(value)
    except (ValueError, TypeError):
        return None
    return number


def dictionary(workbook_path: Path) -> dict[int, dict[str, Any]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb["データ一覧"]
    rows: dict[int, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        idx, variable, question, content, _desc, _num, label = row[:7]
        if isinstance(idx, int):
            rows[idx] = {
                "column": idx,
                "variable": clean(variable),
                "question": clean(question),
                "content": clean(content),
                "label": clean(label),
            }
    wb.close()
    return rows


def sheet_rows(workbook_path: Path, sheet: str) -> list[tuple[Any, ...]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    return rows


def counter_for_columns(rows: list[tuple[Any, ...]], columns: list[int]) -> dict[int, Counter[int]]:
    counters: dict[int, Counter[int]] = {col: Counter() for col in columns}
    for row in rows:
        for col in columns:
            value = numeric(row[col - 1] if col - 1 < len(row) else None)
            if value is not None:
                counters[col][value] += 1
    return counters


def a_burden_profile(rows: list[tuple[Any, ...]], meta: dict[int, dict[str, Any]]) -> list[dict[str, Any]]:
    counters = counter_for_columns(rows, A_BURDEN_COLUMNS)
    profiles: list[dict[str, Any]] = []
    for col in A_BURDEN_COLUMNS:
        counts = counters[col]
        usable = sum(counts.values())
        item = meta[col]["content"]
        profiles.append(
            {
                "column": col,
                "item": item,
                "usable_records": usable,
                "implemented_high_burden": counts[1],
                "implemented_some_burden": counts[2],
                "implemented_no_burden": counts[3],
                "not_implemented_expected_high_burden": counts[4],
                "not_implemented_expected_some_burden": counts[5],
                "not_implemented_expected_no_burden": counts[6],
                "implemented_any": counts[1] + counts[2] + counts[3],
                "implemented_with_any_burden": counts[1] + counts[2],
                "not_implemented_expected_any_burden": counts[4] + counts[5],
                "burden_or_expected_burden_any": counts[1] + counts[2] + counts[4] + counts[5],
                "burden_or_expected_burden_rate": round((counts[1] + counts[2] + counts[4] + counts[5]) / usable, 4) if usable else None,
                "falcon_routes": A_ITEM_ROUTES.get(item, ["QR-03-worksite-contact-and-mobility"]),
                "interpretation_boundary": "establishment-side burden perception or expected-burden signal only; no actual burden validity or guidance claim",
            }
        )
    return sorted(profiles, key=lambda item: (item["burden_or_expected_burden_any"], item["implemented_with_any_burden"]), reverse=True)


def a_advice_profile(rows: list[tuple[Any, ...]], meta: dict[int, dict[str, Any]]) -> list[dict[str, Any]]:
    counters = counter_for_columns(rows, A_ADVICE_COLUMNS)
    profiles: list[dict[str, Any]] = []
    for col in A_ADVICE_COLUMNS:
        counts = counters[col]
        usable = counts[1] + counts[2]
        item = meta[col]["content"]
        profiles.append(
            {
                "column": col,
                "item": item,
                "usable_records": usable,
                "advice_or_external_use_present": counts[1],
                "advice_or_external_use_absent": counts[2],
                "advice_or_external_use_rate": round(counts[1] / usable, 4) if usable else None,
                "falcon_routes": A_ITEM_ROUTES.get(item, ["QR-06-disclosure-boundary-and-mutual-translation"]),
                "interpretation_boundary": "establishment-side external advice/use signal only; no source/support validity judgment",
            }
        )
    return sorted(profiles, key=lambda item: item["advice_or_external_use_present"], reverse=True)


def likert_agreement_profile(rows: list[tuple[Any, ...]], columns: list[int], meta: dict[int, dict[str, Any]], profile_kind: str) -> list[dict[str, Any]]:
    counters = counter_for_columns(rows, columns)
    profiles: list[dict[str, Any]] = []
    for col in columns:
        counts = counters[col]
        usable = sum(counts.values())
        agree = counts[1] + counts[2]
        disagree = counts[4] + counts[5]
        profiles.append(
            {
                "column": col,
                "item": meta[col]["content"],
                "profile_kind": profile_kind,
                "usable_records": usable,
                "agree_or_tend_agree": agree,
                "neutral": counts[3],
                "disagree_or_tend_disagree": disagree,
                "agree_or_tend_agree_rate": round(agree / usable, 4) if usable else None,
                "falcon_routes": ["QR-07-quality-career-and-value-translation", "QR-06-disclosure-boundary-and-mutual-translation"],
                "interpretation_boundary": "establishment-side belief/perception signal only; no employer validity or current guidance claim",
            }
        )
    return sorted(profiles, key=lambda item: item["agree_or_tend_agree"], reverse=True)


def c_condition_window_profile(rows: list[tuple[Any, ...]], meta: dict[int, dict[str, Any]]) -> dict[str, Any]:
    counters = counter_for_columns(rows, C_MAIN_CONDITION_COLUMNS + C_RESPONSE_METHOD_COLUMNS)
    named: dict[str, Any] = {}
    for col in C_MAIN_CONDITION_COLUMNS + C_RESPONSE_METHOD_COLUMNS:
        named[f"{col}:{meta[col]['content'] or meta[col]['variable']}"] = dict(sorted(counters[col].items()))
    return {
        "status": "condition_window_evidence_layer_only",
        "counts_by_column": named,
        "interpretation_boundary": "condition and response-method windows only; no disease/disability-to-support lookup and no worker judgment",
    }


def c_work_content_profile(rows: list[tuple[Any, ...]], meta: dict[int, dict[str, Any]]) -> list[dict[str, Any]]:
    counters = counter_for_columns(rows, C_WORK_CONTENT_COLUMNS)
    profiles: list[dict[str, Any]] = []
    for col in C_WORK_CONTENT_COLUMNS:
        counts = counters[col]
        usable = counts[1] + counts[2] + counts[3]
        yes = counts[1]
        ambiguous = counts[2]
        no = counts[3]
        profiles.append(
            {
                "column": col,
                "item": meta[col]["content"],
                "usable_records": usable,
                "yes": yes,
                "ambiguous": ambiguous,
                "no": no,
                "yes_rate": round(yes / usable, 4) if usable else None,
                "falcon_routes": ["QR-03-worksite-contact-and-mobility", "QR-07-quality-career-and-value-translation"],
                "interpretation_boundary": "worker-side work target signal only; no capacity or suitability judgment",
            }
        )
    return profiles


def c_support_need_profile(rows: list[tuple[Any, ...]], meta: dict[int, dict[str, Any]]) -> list[dict[str, Any]]:
    counters = counter_for_columns(rows, C_SUPPORT_COLUMNS)
    profiles: list[dict[str, Any]] = []
    for group, cols in SUPPORT_GROUPS.items():
        cell_counts: Counter[int] = Counter()
        record_counts: Counter[str] = Counter()
        for row in rows:
            values = [numeric(row[col - 1] if col - 1 < len(row) else None) for col in cols]
            values = [value for value in values if value is not None]
            if not values:
                continue
            cell_counts.update(values)
            if any(value == 1 for value in values):
                record_counts["any_strong_need_or_usefulness"] += 1
            if any(value in {1, 2} for value in values):
                record_counts["any_need_or_usefulness"] += 1
            if all(value in {4, 5} for value in values):
                record_counts["all_unneeded_or_adverse"] += 1
            if any(value == 5 for value in values):
                record_counts["any_adverse_or_counter_usefulness"] += 1
            record_counts["usable_records"] += 1
        usable = record_counts["usable_records"]
        profiles.append(
            {
                "category": group,
                "item_count": len(cols),
                "usable_records": usable,
                "cell_counts": dict(sorted(cell_counts.items())),
                "record_counts": dict(sorted(record_counts.items())),
                "any_need_or_usefulness_rate": round(record_counts["any_need_or_usefulness"] / usable, 4) if usable else None,
                "any_strong_need_or_usefulness_rate": round(record_counts["any_strong_need_or_usefulness"] / usable, 4) if usable else None,
                "any_adverse_or_counter_usefulness_rate": round(record_counts["any_adverse_or_counter_usefulness"] / usable, 4) if usable else None,
                "falcon_routes": SUPPORT_GROUP_ROUTES[group],
                "freedom_axis": SUPPORT_GROUP_FREEDOM[group],
                "interpretation_boundary": "worker-side need/usefulness signal only; no support adequacy or intervention recommendation",
            }
        )
    return sorted(profiles, key=lambda item: item["any_need_or_usefulness_rate"] or 0, reverse=True)


def c_satisfaction_profile(rows: list[tuple[Any, ...]]) -> dict[str, Any]:
    counts = counter_for_columns(rows, [C_SATISFACTION_COLUMN])[C_SATISFACTION_COLUMN]
    usable = sum(counts.values())
    satisfied = counts[1] + counts[2]
    dissatisfied = counts[4] + counts[5]
    return {
        "column": C_SATISFACTION_COLUMN,
        "usable_records": usable,
        "very_satisfied": counts[1],
        "satisfied": counts[2],
        "neutral": counts[3],
        "not_satisfied": counts[4],
        "dissatisfied": counts[5],
        "satisfied_or_very_satisfied": satisfied,
        "dissatisfied_or_not_satisfied": dissatisfied,
        "satisfied_or_very_satisfied_rate": round(satisfied / usable, 4) if usable else None,
        "dissatisfied_or_not_satisfied_rate": round(dissatisfied / usable, 4) if usable else None,
        "falcon_routes": ["QR-07-quality-career-and-value-translation"],
        "interpretation_boundary": "participation-quality modifier only; not proof of support adequacy or employment quality",
    }


def build() -> dict[str, Any]:
    a_meta = dictionary(SOURCE_DIR / "A_data.xlsx")
    c_meta = dictionary(SOURCE_DIR / "C_data.xlsx")
    a_rows = sheet_rows(SOURCE_DIR / "A_data.xlsx", "A全て")
    c_rows = sheet_rows(SOURCE_DIR / "C_data.xlsx", "C全て")

    a_profile = {
        "source_table": "A",
        "row_count": len(a_rows),
        "profile_kind": "establishment_hr_labor_single_perspective",
        "burden_profiles": a_burden_profile(a_rows, a_meta),
        "external_advice_use_profiles": a_advice_profile(a_rows, a_meta),
        "employment_reason_profiles": likert_agreement_profile(a_rows, A_EMPLOYMENT_REASON_COLUMNS, a_meta, "employment_reason"),
        "employment_challenge_profiles": likert_agreement_profile(a_rows, A_EMPLOYMENT_CHALLENGE_COLUMNS, a_meta, "employment_challenge"),
        "network_use": [
            "establishment-side burden and expected-burden window",
            "external advice / consultation / outsourcing connection window",
            "employment reason and challenge perception window",
            "do not use as employer validity or current guidance",
        ],
    }

    c_profile = {
        "source_table": "C",
        "row_count": len(c_rows),
        "profile_kind": "worker_single_perspective",
        "condition_window_profile": c_condition_window_profile(c_rows, c_meta),
        "work_content_target_profile": c_work_content_profile(c_rows, c_meta),
        "support_need_usefulness_profiles": c_support_need_profile(c_rows, c_meta),
        "satisfaction_profile": c_satisfaction_profile(c_rows),
        "network_use": [
            "worker-side condition window and response-method context",
            "worker-side work target and work-content window",
            "worker-side support need/usefulness window",
            "job satisfaction as participation-quality modifier",
            "do not use diagnosis/disability labels as support lookup",
        ],
    }

    return {
        "dataset_id": DATASET_ID,
        "profile_id": "2001_ABC_survey_A_C_single_perspective_structure_profiles_v0_2026_05_22",
        "status": "aggregate_no_text_unreviewed_no_promotion",
        "lane": "Falcon Lab",
        "source_content_exported": False,
        "narrative_content_included": False,
        "row_level_ids_exported": False,
        "A_establishment_profile": a_profile,
        "C_worker_profile": c_profile,
        "network_boundary": [
            "single-perspective evidence layer only",
            "no support validity judgment",
            "no source validity judgment",
            "no disease/disability-to-support lookup",
            "no current policy or workplace guidance claim",
            "no candidate_pattern or reviewed knowledge promotion",
        ],
    }


def write_markdown(data: dict[str, Any]) -> None:
    a = data["A_establishment_profile"]
    c = data["C_worker_profile"]
    lines = [
        "# 2001 ABC Survey A/C Single-Perspective Structure Profiles",
        "",
        "作成日: 2026-05-22",
        "Lane: Falcon Lab",
        "状態: aggregate profile / no narrative text / 未レビュー / 昇格なし",
        "本文引用: なし",
        "",
        "## Position",
        "",
        "A票とC票は、B/C紐付けの補助ではなく、それぞれ独立した専門知識ネットワーク上の観測窓として扱う。",
        "",
        "- A票: 事業所・人事労務側の負担感、外部助言利用、雇用理由/課題の窓。",
        "- C票: 本人側の条件窓、仕事対象、支援必要/有用性、満足度の窓。",
        "",
        "## A票: Burden / Expected Burden",
        "",
        "| item | usable | burden or expected burden | rate | routes |",
        "|---|---:|---:|---:|---|",
    ]
    for item in a["burden_profiles"]:
        routes = ", ".join(f"`{route}`" for route in item["falcon_routes"])
        lines.append(
            f"| {item['item']} | {item['usable_records']} | {item['burden_or_expected_burden_any']} | "
            f"{item['burden_or_expected_burden_rate']} | {routes} |"
        )

    lines.extend(["", "## A票: External Advice / Consultation Use", "", "| item | usable | use present | rate | routes |", "|---|---:|---:|---:|---|"])
    for item in a["external_advice_use_profiles"]:
        routes = ", ".join(f"`{route}`" for route in item["falcon_routes"])
        lines.append(
            f"| {item['item']} | {item['usable_records']} | {item['advice_or_external_use_present']} | "
            f"{item['advice_or_external_use_rate']} | {routes} |"
        )

    lines.extend(["", "## A票: Employment Reasons / Challenges", "", "| kind | item | usable | agree/tend agree | rate |", "|---|---|---:|---:|---:|"])
    for item in a["employment_reason_profiles"] + a["employment_challenge_profiles"]:
        lines.append(
            f"| `{item['profile_kind']}` | {item['item']} | {item['usable_records']} | "
            f"{item['agree_or_tend_agree']} | {item['agree_or_tend_agree_rate']} |"
        )

    lines.extend(["", "## C票: Worker Support Need / Usefulness", "", "| category | usable | any need/usefulness | rate | strong need/usefulness rate | freedom axis | routes |", "|---|---:|---:|---:|---:|---|---|"])
    for item in c["support_need_usefulness_profiles"]:
        routes = ", ".join(f"`{route}`" for route in item["falcon_routes"])
        counts = item["record_counts"]
        lines.append(
            f"| `{item['category']}` | {item['usable_records']} | {counts.get('any_need_or_usefulness', 0)} | "
            f"{item['any_need_or_usefulness_rate']} | {item['any_strong_need_or_usefulness_rate']} | "
            f"`{item['freedom_axis']}` | {routes} |"
        )

    sat = c["satisfaction_profile"]
    lines.extend(
        [
            "",
            "## C票: Satisfaction As Participation-Quality Modifier",
            "",
            f"- usable records: {sat['usable_records']}",
            f"- satisfied or very satisfied: {sat['satisfied_or_very_satisfied']} ({sat['satisfied_or_very_satisfied_rate']})",
            f"- dissatisfied or not satisfied: {sat['dissatisfied_or_not_satisfied']} ({sat['dissatisfied_or_not_satisfied_rate']})",
            "",
            "## C票: Work Target Signals",
            "",
            "| item | usable | yes | yes rate |",
            "|---|---:|---:|---:|",
        ]
    )
    for item in c["work_content_target_profile"]:
        lines.append(f"| {item['item']} | {item['usable_records']} | {item['yes']} | {item['yes_rate']} |")

    lines.extend(
        [
            "",
            "## Boundary",
            "",
            "- A票は事業所側の認識窓であり、負担の正当性や現在の事業主助言ではない。",
            "- C票は本人側の認識窓であり、支援妥当性や就労能力判断ではない。",
            "- 条件窓・診断名・障害種類から配慮を引かない。",
            "- このprofileはA/B/C三者統合カードの材料であり、知識昇格ではない。",
            "",
        ]
    )
    OUTPUT_MD.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    DERIVED_DIR.mkdir(parents=True, exist_ok=True)
    data = build()
    OUTPUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_markdown(data)
    print(
        json.dumps(
            {
                "A_rows": data["A_establishment_profile"]["row_count"],
                "C_rows": data["C_worker_profile"]["row_count"],
                "A_burden_items": len(data["A_establishment_profile"]["burden_profiles"]),
                "C_support_groups": len(data["C_worker_profile"]["support_need_usefulness_profiles"]),
                "output": str(OUTPUT_JSON.relative_to(REPO_ROOT)),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
