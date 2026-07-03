#!/usr/bin/env python3
"""Build redacted narrative staging for the 2001 ABC survey.

This is a local preprocessing step for SCIMA/FCHMA analysis. It reads raw
open-text cells from data/original_secure, writes redacted staging data under
data/staging/anonymized, and writes aggregate no-text summaries under
references/derived.
"""

from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from redact_free_text_units import QA_PATTERNS, redact_text, sha256_text  # noqa: E402


REPO_ROOT = SCRIPT_DIR.parents[1]
DATASET_ID = "2001_ABC_survey"
SOURCE_DIR = REPO_ROOT / "data/original_secure/structured/2001_ABC_survey"
STAGING_DIR = REPO_ROOT / "data/staging/anonymized/2001_ABC_survey/v0"
DERIVED_DIR = REPO_ROOT / "references/derived/scima-fchma/2001-abc-survey-v0-2026-05-22"


TABLES: dict[str, dict[str, Any]] = {
    "A": {
        "workbook": SOURCE_DIR / "A_data.xlsx",
        "sheet": "A全て",
        "perspective": "establishment_hr_labor",
        "id_columns": {"establishment_id": 1},
    },
    "B": {
        "workbook": SOURCE_DIR / "B_data.xlsx",
        "sheet": "B全て",
        "perspective": "supervisor_workplace",
        "id_columns": {"establishment_id": 1, "target_id": 2},
    },
    "C": {
        "workbook": SOURCE_DIR / "C_data.xlsx",
        "sheet": "C全て",
        "perspective": "worker",
        "id_columns": {"establishment_id": 1, "target_id": 2},
    },
}


ROLE_RULES: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"具体的仕事内容|実際の作業内容|仕事内容"), "work_content_narrative"),
    (re.compile(r"相談先|相談内容|外部支援機関"), "external_support_advice_narrative"),
    (re.compile(r"自由記述|その他の課題|支援上のニーズ|支援や制度"), "general_open_narrative"),
    (re.compile(r"その他の有効な配慮|その他の配慮|支援方法"), "support_practice_narrative"),
    (re.compile(r"事業の取り扱い品|定着推進活動状況|障害者雇用留意事項"), "A_added_establishment_context_narrative"),
    (re.compile(r"その他の雇用理由|その他の雇用課題"), "A_employment_reason_challenge_narrative"),
    (re.compile(r"その他具体的に|その他の障害|その他"), "condition_or_other_specification_text"),
    (re.compile(r"具体的診断名"), "sensitive_health_condition_text"),
]

OPEN_TEXT_HINT = re.compile(
    r"その他|具体|自由記述|相談|内容|診断名|取り扱い品|留意事項|定着推進|作業内容|仕事内容|支援方法|外部支援"
)


SIGNAL_RULES: dict[str, list[str]] = {
    "work_content_task": ["作業", "業務", "仕事", "職務", "製造", "事務", "販売", "清掃", "入力", "検査", "接客"],
    "worksite_contact_environment": ["設備", "機器", "通路", "階段", "トイレ", "駐車", "配置", "作業台", "表示", "音", "照明"],
    "information_communication": ["説明", "連絡", "会議", "会話", "手話", "筆談", "メール", "ファックス", "伝達", "相談"],
    "health_time": ["通院", "治療", "服薬", "体調", "疲労", "休憩", "勤務時間", "残業", "健康", "診療"],
    "support_practice": ["配慮", "支援", "補助", "介助", "指導", "相談員", "援助", "訓練", "調整"],
    "external_support_system": ["病院", "学校", "福祉", "施設", "職安", "ハローワーク", "医師", "家族", "支援機関"],
    "participation_quality": ["満足", "不満", "やりがい", "人間関係", "上司", "同僚", "職場", "偏見", "理解"],
    "burden_management": ["負担", "課題", "困難", "留意", "管理", "定着", "問題", "安全", "責任"],
    "commuting_daily_living": ["通勤", "送迎", "住居", "寮", "生活", "買い物", "自立", "家族"],
    "sensitive_health_condition": ["診断", "障害名", "病名", "疾患"],
}


def clean_cell_text(value: Any) -> str:
    if value is None or isinstance(value, bool):
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return ""
    if text.replace(".", "", 1).isdigit():
        return ""
    return text


def clean_identifier(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def classify_role(variable: str, question: str, content: str) -> str:
    joined = " ".join([variable, question, content])
    for pattern, role in ROLE_RULES:
        if pattern.search(joined):
            return role
    return "open_text_needs_review"


def split_units(text: str) -> list[str]:
    pieces: list[str] = []
    for line in re.split(r"\n+", text):
        line = line.strip()
        if not line:
            continue
        if len(line) <= 450:
            pieces.append(line)
            continue
        current = ""
        for part in re.split(r"(?<=[。！？!?])", line):
            part = part.strip()
            if not part:
                continue
            if len(current) + len(part) > 450 and current:
                pieces.append(current)
                current = part
            else:
                current += part
        if current:
            pieces.append(current)
    return pieces


def dictionary_rows(workbook_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb["データ一覧"]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        idx, variable, question, content, _desc, _num, label = row[:7]
        if not isinstance(idx, int):
            continue
        rows.append(
            {
                "column": idx,
                "variable": clean_cell_text(variable),
                "question": clean_cell_text(question),
                "content": clean_cell_text(content),
                "label": clean_cell_text(label),
            }
        )
    wb.close()
    return rows


def narrative_fields(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    fields: list[dict[str, Any]] = []
    for row in rows:
        joined = " ".join([row["variable"], row["question"], row["content"]])
        if row["label"] or not OPEN_TEXT_HINT.search(joined):
            continue
        field = dict(row)
        field["field_role"] = classify_role(row["variable"], row["question"], row["content"])
        fields.append(field)
    return fields


def linkage_hash(table: str, row_values: tuple[Any, ...], id_columns: dict[str, int]) -> str:
    parts = [DATASET_ID, table]
    for name, col in sorted(id_columns.items()):
        value = clean_identifier(row_values[col - 1] if col - 1 < len(row_values) else "")
        parts.append(f"{name}:{value}")
    return sha256_text("|".join(parts))[:24]


def cross_table_hashes(row_values: tuple[Any, ...], id_columns: dict[str, int]) -> dict[str, str | None]:
    establishment_col = id_columns.get("establishment_id")
    target_col = id_columns.get("target_id")
    establishment_id = clean_identifier(row_values[establishment_col - 1]) if establishment_col else ""
    target_id = clean_identifier(row_values[target_col - 1]) if target_col else ""
    establishment_hash = sha256_text(f"{DATASET_ID}|establishment:{establishment_id}")[:24] if establishment_id else None
    pair_hash = (
        sha256_text(f"{DATASET_ID}|establishment:{establishment_id}|target:{target_id}")[:24]
        if establishment_id and target_id
        else None
    )
    return {
        "establishment_hash": establishment_hash,
        "pair_hash": pair_hash,
    }


def unit_id(table: str, link_hash: str, column: int, unit_index: int, original_hash: str) -> str:
    return f"{DATASET_ID}:{table}:{link_hash}:c{column}:u{unit_index}:{original_hash[:12]}"


def signal_flags(text: str) -> list[str]:
    flags: list[str] = []
    for signal, keywords in SIGNAL_RULES.items():
        if any(keyword in text for keyword in keywords):
            flags.append(signal)
    return flags


def relative(path: Path) -> str:
    return str(path.relative_to(REPO_ROOT))


def main() -> None:
    STAGING_DIR.mkdir(parents=True, exist_ok=True)
    DERIVED_DIR.mkdir(parents=True, exist_ok=True)

    output_jsonl = STAGING_DIR / "narrative_units.redacted.jsonl"
    flags_jsonl = STAGING_DIR / "narrative_redaction_review_flags.jsonl"
    manifest_path = STAGING_DIR / "narrative_redaction_manifest.json"
    summary_json = DERIVED_DIR / "2001-abc-survey-narrative-staging-summary-v0-2026-05-22.json"
    summary_md = DERIVED_DIR / "2001-abc-survey-narrative-staging-summary-v0-2026-05-22.md"

    total_units = 0
    changed_units = 0
    source_cells = 0
    redaction_counts: Counter[str] = Counter()
    residual_counts: Counter[str] = Counter()
    role_counts: Counter[str] = Counter()
    signal_counts: Counter[str] = Counter()
    table_counts: Counter[str] = Counter()
    table_role_counts: dict[str, Counter[str]] = defaultdict(Counter)
    table_signal_counts: dict[str, Counter[str]] = defaultdict(Counter)
    flags_written = 0

    with output_jsonl.open("w", encoding="utf-8") as out, flags_jsonl.open("w", encoding="utf-8") as flags_out:
        for table, config in TABLES.items():
            rows = dictionary_rows(config["workbook"])
            fields = narrative_fields(rows)
            field_by_col = {field["column"]: field for field in fields}
            wb = load_workbook(config["workbook"], read_only=False, data_only=True)
            ws = wb[config["sheet"]]
            target_cols = sorted(field_by_col)

            for row_index, row_values in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
                link_hash = linkage_hash(table, row_values, config["id_columns"])
                cross_hashes = cross_table_hashes(row_values, config["id_columns"])
                for col in target_cols:
                    raw_text = clean_cell_text(row_values[col - 1] if col - 1 < len(row_values) else "")
                    if not raw_text:
                        continue
                    source_cells += 1
                    field = field_by_col[col]
                    units = split_units(raw_text)
                    for local_index, raw_unit in enumerate(units, start=1):
                        original_hash = sha256_text(raw_unit)
                        redacted, counts = redact_text(raw_unit)
                        if redacted != raw_unit:
                            changed_units += 1
                        redaction_counts.update(counts)
                        residual = {
                            name: len(pattern.findall(redacted))
                            for name, pattern in QA_PATTERNS.items()
                        }
                        residual = {name: count for name, count in residual.items() if count}
                        residual_counts.update(residual)

                        total_units += 1
                        role = field["field_role"]
                        signals = signal_flags(redacted)
                        role_counts[role] += 1
                        table_counts[table] += 1
                        table_role_counts[table][role] += 1
                        signal_counts.update(signals)
                        table_signal_counts[table].update(signals)

                        record = {
                            "dataset_id": DATASET_ID,
                            "source_table": table,
                            "perspective": config["perspective"],
                            "linkage_key_hash": link_hash,
                            "establishment_hash": cross_hashes["establishment_hash"],
                            "pair_hash": cross_hashes["pair_hash"],
                            "unit_id": unit_id(table, link_hash, col, local_index, original_hash),
                            "row_ref_hash": sha256_text(f"{DATASET_ID}|{table}|row:{row_index}")[:24],
                            "column": col,
                            "variable": field["variable"],
                            "content": field["content"],
                            "field_role": role,
                            "unit_index_in_cell": local_index,
                            "source_char_count": len(raw_unit),
                            "original_content_hash": original_hash,
                            "redacted_unit_text": redacted,
                            "redacted_content_hash": sha256_text(redacted),
                            "redacted_char_count": len(redacted),
                            "redaction_types": sorted(counts.keys()),
                            "redaction_count": sum(counts.values()),
                            "deterministic_scima_fchma_signal_flags": signals,
                            "sensitive_health_text": role == "sensitive_health_condition_text",
                            "source_content_exported": False,
                            "redaction_status": "local_deterministic_redaction_not_human_reviewed",
                            "review_status": "unreviewed_staging",
                        }
                        out.write(json.dumps(record, ensure_ascii=False) + "\n")

                        if residual or role == "sensitive_health_condition_text":
                            flags_written += 1
                            flags_out.write(
                                json.dumps(
                                    {
                                        "dataset_id": DATASET_ID,
                                        "source_table": table,
                                        "unit_id": record["unit_id"],
                                        "field_role": role,
                                        "residual_pattern_counts": residual,
                                        "review_reasons": [
                                            *(
                                                ["residual_obvious_identifier_pattern_after_redaction"]
                                                if residual
                                                else []
                                            ),
                                            *(
                                                ["sensitive_health_condition_text"]
                                                if role == "sensitive_health_condition_text"
                                                else []
                                            ),
                                        ],
                                    },
                                    ensure_ascii=False,
                                )
                                + "\n"
                            )
            wb.close()

    manifest = {
        "dataset_id": DATASET_ID,
        "status": "local_deterministic_redaction_not_human_reviewed",
        "output": relative(output_jsonl),
        "review_flags": relative(flags_jsonl),
        "source_content_exported": False,
        "anonymized_content_public_safe": False,
        "records": total_units,
        "source_nonempty_cells": source_cells,
        "records_changed": changed_units,
        "redaction_type_counts": dict(sorted(redaction_counts.items())),
        "residual_obvious_identifier_pattern_counts": dict(sorted(residual_counts.items())),
        "review_flagged_records": flags_written,
        "notes": [
            "Raw narrative text was read locally and not written to references.",
            "Redacted text remains sensitive staging data and is not public-safe.",
            "Sensitive health-condition fields are flagged for separate review.",
            "Deterministic signal flags are only routing hints for SCIMA/FCHMA, not analysis conclusions.",
        ],
    }
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    summary = {
        "dataset_id": DATASET_ID,
        "summary_id": "2001_ABC_survey_narrative_staging_summary_v0_2026_05_22",
        "status": "aggregate_no_text_summary_from_redacted_staging",
        "source_manifest": relative(manifest_path),
        "source_content_exported": False,
        "anonymized_content_exported_to_references": False,
        "total_units": total_units,
        "source_nonempty_cells": source_cells,
        "table_unit_counts": dict(sorted(table_counts.items())),
        "field_role_counts": dict(sorted(role_counts.items())),
        "signal_counts": dict(sorted(signal_counts.items())),
        "table_role_counts": {table: dict(sorted(counts.items())) for table, counts in sorted(table_role_counts.items())},
        "table_signal_counts": {table: dict(sorted(counts.items())) for table, counts in sorted(table_signal_counts.items())},
        "review_flagged_records": flags_written,
        "first_card_families": [
            "A_establishment_narrative_context_card",
            "B_work_content_contact_narrative_card",
            "B_support_practice_narrative_card",
            "C_worker_work_content_narrative_card",
            "C_worker_need_policy_narrative_card",
        ],
        "boundary": [
            "no raw quotation",
            "no row-level text in references",
            "no support validity judgment",
            "no diagnosis-to-support lookup",
            "no knowledge promotion",
        ],
    }
    summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    lines = [
        "# 2001 ABC Survey Narrative Staging Summary",
        "",
        "作成日: 2026-05-22",
        "Lane: Falcon Lab",
        "状態: aggregate no-text summary / redacted staging generated / 未レビュー / 統合なし / 昇格なし",
        "本文引用: なし",
        "row-level回答本文のreferences外部化: なし",
        "",
        "## What Was Generated",
        "",
        f"- redacted staging units: `{relative(output_jsonl)}`",
        f"- redaction manifest: `{relative(manifest_path)}`",
        f"- redaction review flags: `{relative(flags_jsonl)}`",
        f"- aggregate JSON summary: `{relative(summary_json)}`",
        "",
        "## Scale",
        "",
        f"- source nonempty narrative cells: {source_cells}",
        f"- redacted narrative units: {total_units}",
        f"- units changed by deterministic redaction: {changed_units}",
        f"- review flagged units: {flags_written}",
        "",
        "## Units By Table",
        "",
        "| table | units |",
        "|---|---:|",
    ]
    for table, count in sorted(table_counts.items()):
        lines.append(f"| `{table}` | {count} |")
    lines.extend(["", "## Field Roles", "", "| role | units |", "|---|---:|"])
    for role, count in sorted(role_counts.items()):
        lines.append(f"| `{role}` | {count} |")
    lines.extend(["", "## Deterministic SCIMA/FCHMA Routing Signals", "", "| signal | units |", "|---|---:|"])
    for signal, count in sorted(signal_counts.items(), key=lambda item: (-item[1], item[0])):
        lines.append(f"| `{signal}` | {count} |")
    lines.extend(
        [
            "",
            "## Boundary",
            "",
            "- Redacted text remains staging data, not public-safe content.",
            "- Signal flags are routing hints, not SCIMA/FCHMA conclusions.",
            "- Sensitive health-condition text is flagged and must not become a disease-to-support route.",
            "- The next valid step is single-perspective narrative card generation before linked contrast cards.",
            "",
        ]
    )
    summary_md.write_text("\n".join(lines), encoding="utf-8")

    print(
        json.dumps(
            {
                "records": total_units,
                "source_nonempty_cells": source_cells,
                "records_changed": changed_units,
                "review_flagged_records": flags_written,
                "output": relative(output_jsonl),
                "manifest": relative(manifest_path),
                "summary": relative(summary_json),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
