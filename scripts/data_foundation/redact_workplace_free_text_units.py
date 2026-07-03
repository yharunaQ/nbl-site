#!/usr/bin/env python3
"""Extract and redact workplace-survey free-text columns without raw text output."""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from redact_free_text_units import QA_PATTERNS, redact_text, sha256_text  # noqa: E402


REPO_ROOT = SCRIPT_DIR.parents[1]


DATASETS: dict[str, dict[str, Any]] = {
    "nanbyo_workplace_2022_2023": {
        "workbook": REPO_ROOT / "data/original_secure/structured/nanbyo_workplace_2022_2023/難病職場調査.xlsx",
        "sheet": "ローデータ数値化",
        "header_rows": [1, 2],
        "data_start_row": 3,
        "output_dir": REPO_ROOT / "data/staging/anonymized/nanbyo_workplace_2022_2023/v0",
        "columns": [
            {"column_ref": "M", "reason": "industry other text"},
            {"column_ref": "AD", "reason": "other recognition situation text"},
            {"column_ref": "BL", "reason": "disease name text for most-support-needed case"},
            {"column_ref": "CT", "reason": "disease name text for least-support-needed case"},
            {"column_ref": "ET", "reason": "free text opinion / issues / concerns"},
        ],
        "excluded_identifiable_columns": [],
    },
    "nanbyo_workplace_2022_2023_web_raw0324": {
        "workbook": REPO_ROOT / "data/original_secure/structured/nanbyo_workplace_2022_2023/事業所調査（ローデータ0324のみ抽出）.xlsx",
        "sheet": "事業所調査（回答者ローデータ0324）",
        "header_rows": [1, 2, 3],
        "data_start_row": 4,
        "output_dir": REPO_ROOT / "data/staging/anonymized/nanbyo_workplace_2022_2023_web_raw0324/v0",
        "columns": [
            {"column_ref": "O", "reason": "other text / free-text candidate"},
            {"column_ref": "AF", "reason": "other text / free-text candidate"},
            {"column_ref": "BN", "reason": "disease name / sensitive text candidate"},
            {"column_ref": "CV", "reason": "disease name / sensitive text candidate"},
            {"column_ref": "EV", "reason": "free text opinion / issues / concerns"},
        ],
        "excluded_identifiable_columns": ["D", "E"],
    },
}


def clean_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, (int, float)):
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return ""
    if text.replace(".", "", 1).isdigit():
        return ""
    return text


def header_label(ws: Any, column_ref: str, header_rows: list[int]) -> str:
    col_idx = column_index_from_string(column_ref)
    parts: list[str] = []
    for row_idx in header_rows:
        value = ws.cell(row=row_idx, column=col_idx).value
        text = clean_cell_text(value)
        if text and text not in parts:
            parts.append(text)
    return " / ".join(parts)


def iter_redacted_records(dataset_id: str, config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    wb = load_workbook(config["workbook"], read_only=False, data_only=True)
    ws = wb[config["sheet"]]

    total_rows_scanned = 0
    total_records = 0
    changed = 0
    type_counts: Counter[str] = Counter()
    residual_counts: Counter[str] = Counter()
    flagged = 0
    by_column: dict[str, Counter[str]] = defaultdict(Counter)
    records: list[dict[str, Any]] = []
    flags: list[dict[str, Any]] = []

    labels = {
        col["column_ref"]: header_label(ws, col["column_ref"], config["header_rows"])
        for col in config["columns"]
    }

    target_columns = [
        (col, column_index_from_string(col["column_ref"]))
        for col in config["columns"]
    ]

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=config["data_start_row"], max_row=ws.max_row, values_only=True),
        start=config["data_start_row"],
    ):
        total_rows_scanned += 1
        row_unit_count = 0
        for col, col_idx in target_columns:
            column_ref = col["column_ref"]
            raw_text = clean_cell_text(row[col_idx - 1] if col_idx - 1 < len(row) else None)
            if not raw_text:
                continue

            total_records += 1
            row_unit_count += 1
            redacted_text, counts = redact_text(raw_text)
            type_counts.update(counts)
            by_column[column_ref].update(counts)
            if redacted_text != raw_text:
                changed += 1

            residual = {
                name: len(pattern.findall(redacted_text))
                for name, pattern in QA_PATTERNS.items()
            }
            residual = {name: count for name, count in residual.items() if count}
            residual_counts.update(residual)
            if residual:
                flagged += 1
                flags.append({
                    "dataset_id": dataset_id,
                    "row_ref": f"row_{row_idx:04d}",
                    "column_ref": column_ref,
                    "residual_pattern_counts": residual,
                    "review_reason": "residual_obvious_identifier_pattern_after_redaction",
                })

            record = {
                "dataset_id": dataset_id,
                "row_ref": f"row_{row_idx:04d}",
                "respondent_id": f"{dataset_id}:row_{row_idx:04d}",
                "unit_index": row_unit_count,
                "column_ref": column_ref,
                "raw_name": labels[column_ref],
                "column_reason": col["reason"],
                "redacted_unit_text": redacted_text,
                "redaction_types": sorted(counts.keys()),
                "redaction_count": sum(counts.values()),
                "original_content_hash": sha256_text(raw_text),
                "redacted_content_hash": sha256_text(redacted_text),
                "redacted_char_count": len(redacted_text),
                "redaction_status": "local_deterministic_redaction_not_human_reviewed",
            }
            records.append(record)

    wb.close()

    manifest = {
        "dataset_id": dataset_id,
        "input_workbook": str(config["workbook"].relative_to(REPO_ROOT)),
        "sheet": config["sheet"],
        "output": str((config["output_dir"] / "free_text_units.redacted.jsonl").relative_to(REPO_ROOT)),
        "review_flags": str((config["output_dir"] / "redaction_review_flags.jsonl").relative_to(REPO_ROOT)),
        "status": "local_deterministic_redaction_not_human_reviewed",
        "rows_scanned": total_rows_scanned,
        "records": total_records,
        "records_changed": changed,
        "redaction_type_counts": dict(sorted(type_counts.items())),
        "redaction_type_counts_by_column": {
            column_ref: dict(sorted(counter.items()))
            for column_ref, counter in sorted(by_column.items())
        },
        "residual_obvious_identifier_pattern_counts": dict(sorted(residual_counts.items())),
        "residual_flagged_records": flagged,
        "source_raw_text_exported": False,
        "forced_identifiable_columns_exported": False,
        "excluded_identifiable_columns": config["excluded_identifiable_columns"],
        "redacted_text_public_safe": False,
        "notes": [
            "Raw workplace free text was read only locally and was not written to an intermediate artifact.",
            "Forced identifiable-risk columns are excluded from extraction.",
            "Disease-name text remains sensitive health-condition staging data and is not public-safe.",
            "This is preprocessing for SCIMA/FCHMA analysis, not human-reviewed knowledge.",
        ],
    }
    return records, {"manifest": manifest, "flags": flags}


def write_dataset(dataset_id: str) -> dict[str, Any]:
    config = DATASETS[dataset_id]
    config["output_dir"].mkdir(parents=True, exist_ok=True)
    records, meta = iter_redacted_records(dataset_id, config)

    output_jsonl = config["output_dir"] / "free_text_units.redacted.jsonl"
    manifest_path = config["output_dir"] / "redaction_manifest.json"
    flags_path = config["output_dir"] / "redaction_review_flags.jsonl"

    with output_jsonl.open("w", encoding="utf-8") as out:
        for record in records:
            out.write(json.dumps(record, ensure_ascii=False) + "\n")
    manifest_path.write_text(json.dumps(meta["manifest"], ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    with flags_path.open("w", encoding="utf-8") as out:
        for flag in meta["flags"]:
            out.write(json.dumps(flag, ensure_ascii=False) + "\n")

    return {
        "dataset_id": dataset_id,
        "records": meta["manifest"]["records"],
        "records_changed": meta["manifest"]["records_changed"],
        "residual_flagged_records": meta["manifest"]["residual_flagged_records"],
        "output": str(output_jsonl.relative_to(REPO_ROOT)),
        "manifest": str(manifest_path.relative_to(REPO_ROOT)),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--dataset-id",
        choices=[*DATASETS.keys(), "all"],
        default="all",
    )
    args = parser.parse_args()

    dataset_ids = list(DATASETS) if args.dataset_id == "all" else [args.dataset_id]
    summaries = [write_dataset(dataset_id) for dataset_id in dataset_ids]
    print(json.dumps(summaries, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
