#!/usr/bin/env python3
"""Build no-text A/B/C linked structured triad profile for 2001 ABC.

This uses coded structured values only. It exports aggregate cells, not row ids
or narrative text. The profile is for Falcon network enrichment, not support
adequacy judgment.
"""

from __future__ import annotations

import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
DATASET_ID = "2001_ABC_survey"
SOURCE_DIR = REPO_ROOT / "data/original_secure/structured/2001_ABC_survey"
DERIVED_DIR = REPO_ROOT / "references/derived/scima-fchma/2001-abc-survey-v0-2026-05-22"

OUTPUT_JSON = DERIVED_DIR / "2001-abc-survey-A-B-C-linked-structured-triad-profile-v0-2026-05-22.json"
OUTPUT_MD = DERIVED_DIR / "2001-abc-survey-A-B-C-linked-structured-triad-profile-v0-2026-05-22.md"


B_GROUP_COLUMNS = {
    "training_skill_education": list(range(26, 30)),
    "job_task_design_and_performance": list(range(30, 36)),
    "stress_responsibility_safety": list(range(36, 41)),
    "information_communication": list(range(41, 46)),
    "within_workplace_mobility": list(range(46, 53)),
    "commuting": list(range(53, 62)),
    "health_time_self_care_environment": list(range(62, 70)),
    "interpersonal_relations_and_workplace_culture": list(range(70, 76)),
    "emergency_disaster": list(range(76, 79)),
    "offwork_daily_living": list(range(79, 83)),
}

C_GROUP_COLUMNS = {
    "training_skill_education": list(range(44, 48)),
    "job_task_design_and_performance": list(range(48, 54)),
    "stress_responsibility_safety": list(range(54, 59)),
    "information_communication": list(range(59, 64)),
    "within_workplace_mobility": list(range(64, 71)),
    "commuting": list(range(71, 80)),
    "health_time_self_care_environment": list(range(80, 88)),
    "interpersonal_relations_and_workplace_culture": list(range(88, 94)),
    "emergency_disaster": list(range(94, 97)),
    "offwork_daily_living": list(range(97, 102)),
}

A_BURDEN_TO_ADVICE = {burden_col: advice_col for burden_col, advice_col in zip(range(41, 55), range(65, 92, 2))}

TRIAD_CATEGORIES = {
    "worksite_contact_and_equipment": {
        "a_burden_cols": [41, 42, 43, 48, 50],
        "b_groups": ["job_task_design_and_performance", "stress_responsibility_safety", "within_workplace_mobility", "emergency_disaster"],
        "c_groups": ["job_task_design_and_performance", "stress_responsibility_safety", "within_workplace_mobility", "emergency_disaster"],
        "routes": ["QR-03-worksite-contact-and-mobility", "QR-08-diversity-conditioned-same-structure"],
        "freedom_axes": ["body_environment_freedom", "load_quality_safety_freedom", "work_content_freedom"],
    },
    "support_coordination_and_training": {
        "a_burden_cols": [44, 45, 46, 47, 49],
        "b_groups": ["training_skill_education", "job_task_design_and_performance", "interpersonal_relations_and_workplace_culture"],
        "c_groups": ["training_skill_education", "job_task_design_and_performance", "interpersonal_relations_and_workplace_culture"],
        "routes": ["QR-06-disclosure-boundary-and-mutual-translation", "QR-07-quality-career-and-value-translation"],
        "freedom_axes": ["support_coordination_freedom", "learning_variability_freedom", "social_participation_freedom"],
    },
    "information_procedure_and_disclosure": {
        "a_burden_cols": [43, 46, 47],
        "b_groups": ["information_communication", "training_skill_education", "interpersonal_relations_and_workplace_culture"],
        "c_groups": ["information_communication", "training_skill_education", "interpersonal_relations_and_workplace_culture"],
        "routes": ["QR-02-information-work-procedure", "QR-06-disclosure-boundary-and-mutual-translation"],
        "freedom_axes": ["information_translation_freedom", "disclosure_boundary_freedom"],
    },
    "health_time_and_environment": {
        "a_burden_cols": [51],
        "b_groups": ["health_time_self_care_environment", "stress_responsibility_safety"],
        "c_groups": ["health_time_self_care_environment", "stress_responsibility_safety"],
        "routes": ["QR-01-health-time-work-design", "QR-04-life-security-sequencing"],
        "freedom_axes": ["health_time_freedom", "load_quality_safety_freedom"],
    },
    "life_security_and_external_bridge": {
        "a_burden_cols": [52, 53, 54],
        "b_groups": ["commuting", "offwork_daily_living", "health_time_self_care_environment"],
        "c_groups": ["commuting", "offwork_daily_living", "health_time_self_care_environment"],
        "routes": ["QR-04-life-security-sequencing", "QR-05-entry-prework-translation"],
        "freedom_axes": ["life_commuting_freedom", "life_security_freedom", "institutional_connection_freedom"],
    },
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def numeric(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def load_rows(workbook: str, sheet: str) -> list[tuple[Any, ...]]:
    wb = load_workbook(SOURCE_DIR / workbook, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    return rows


def val(row: tuple[Any, ...], col: int) -> int | None:
    if col - 1 >= len(row):
        return None
    return numeric(row[col - 1])


def any_value(row: tuple[Any, ...], cols: list[int], values: set[int]) -> bool:
    return any(val(row, col) in values for col in cols)


def has_any_response(row: tuple[Any, ...], cols: list[int]) -> bool:
    return any(val(row, col) is not None for col in cols)


def group_cols(groups: dict[str, list[int]], names: list[str]) -> list[int]:
    cols: list[int] = []
    for name in names:
        cols.extend(groups[name])
    return cols


def index_a(rows: list[tuple[Any, ...]]) -> dict[str, tuple[Any, ...]]:
    return {clean(row[0]): row for row in rows if clean(row[0])}


def index_pair(rows: list[tuple[Any, ...]]) -> dict[tuple[str, str], tuple[Any, ...]]:
    return {(clean(row[0]), clean(row[1])): row for row in rows if clean(row[0]) and clean(row[1])}


def state_label(flag: bool | None, yes: str, no: str, missing: str) -> str:
    if flag is None:
        return missing
    return yes if flag else no


def triad_state(a_row: tuple[Any, ...], b_row: tuple[Any, ...], c_row: tuple[Any, ...], spec: dict[str, Any]) -> dict[str, Any]:
    a_burden_cols = spec["a_burden_cols"]
    a_advice_cols = [A_BURDEN_TO_ADVICE[col] for col in a_burden_cols]
    b_cols = group_cols(B_GROUP_COLUMNS, spec["b_groups"])
    c_cols = group_cols(C_GROUP_COLUMNS, spec["c_groups"])

    a_burden = any_value(a_row, a_burden_cols, {1, 2, 4, 5}) if has_any_response(a_row, a_burden_cols) else None
    a_advice = any_value(a_row, a_advice_cols, {1}) if has_any_response(a_row, a_advice_cols) else None
    b_support = any_value(b_row, b_cols, {1}) if has_any_response(b_row, b_cols) else None
    b_needed_not_present = any_value(b_row, b_cols, {2}) if has_any_response(b_row, b_cols) else None
    c_need = any_value(c_row, c_cols, {1, 2}) if has_any_response(c_row, c_cols) else None
    c_strong_need = any_value(c_row, c_cols, {1}) if has_any_response(c_row, c_cols) else None

    return {
        "A_burden_or_expected_burden": a_burden,
        "A_external_advice_use": a_advice,
        "B_support_present": b_support,
        "B_supervisor_needed_but_not_present": b_needed_not_present,
        "C_need_or_usefulness": c_need,
        "C_strong_need_or_usefulness": c_strong_need,
        "combo": "|".join(
            [
                state_label(a_burden, "A_burden", "A_no_burden", "A_missing"),
                state_label(a_advice, "A_advice", "A_no_advice", "A_advice_missing"),
                state_label(b_support, "B_support", "B_no_support", "B_missing"),
                state_label(b_needed_not_present, "B_needed_not_present", "B_no_unmet_supervisor_need", "B_unmet_missing"),
                state_label(c_need, "C_need", "C_no_need", "C_missing"),
            ]
        ),
    }


def build_profile() -> dict[str, Any]:
    a = index_a(load_rows("A_data.xlsx", "A全て"))
    b = index_pair(load_rows("B_data.xlsx", "B全て"))
    c = index_pair(load_rows("C_data.xlsx", "C全て"))
    strict_pairs = sorted(set(b) & set(c) & {(est, target) for est, target in b if est in a})

    profiles: list[dict[str, Any]] = []
    for category, spec in TRIAD_CATEGORIES.items():
        counts: Counter[str] = Counter()
        boolean_counts: Counter[str] = Counter()
        for key in strict_pairs:
            state = triad_state(a[key[0]], b[key], c[key], spec)
            counts[state["combo"]] += 1
            for field, value in state.items():
                if field == "combo":
                    continue
                boolean_counts[f"{field}={value}"] += 1

        c_need = boolean_counts["C_need_or_usefulness=True"]
        b_support = boolean_counts["B_support_present=True"]
        a_burden = boolean_counts["A_burden_or_expected_burden=True"]
        a_advice = boolean_counts["A_external_advice_use=True"]
        b_unmet = boolean_counts["B_supervisor_needed_but_not_present=True"]
        profiles.append(
            {
                "category": category,
                "triad_pairs": len(strict_pairs),
                "routes": spec["routes"],
                "freedom_axes": spec["freedom_axes"],
                "A_burden_or_expected_burden": a_burden,
                "A_burden_or_expected_burden_rate": round(a_burden / len(strict_pairs), 4) if strict_pairs else None,
                "A_external_advice_use": a_advice,
                "A_external_advice_use_rate": round(a_advice / len(strict_pairs), 4) if strict_pairs else None,
                "B_support_present": b_support,
                "B_support_present_rate": round(b_support / len(strict_pairs), 4) if strict_pairs else None,
                "B_supervisor_needed_but_not_present": b_unmet,
                "B_supervisor_needed_but_not_present_rate": round(b_unmet / len(strict_pairs), 4) if strict_pairs else None,
                "C_need_or_usefulness": c_need,
                "C_need_or_usefulness_rate": round(c_need / len(strict_pairs), 4) if strict_pairs else None,
                "A_burden_and_C_need_without_B_support": counts_total(
                    counts,
                    required=["A_burden", "C_need", "B_no_support"],
                    forbidden=[],
                ),
                "A_no_burden_but_C_need": counts_total(counts, required=["A_no_burden", "C_need"], forbidden=[]),
                "B_support_and_C_no_need": counts_total(counts, required=["B_support", "C_no_need"], forbidden=[]),
                "A_advice_and_C_need": counts_total(counts, required=["A_advice", "C_need"], forbidden=[]),
                "top_state_combinations": [
                    {"state": state, "count": count, "rate": round(count / len(strict_pairs), 4) if strict_pairs else None}
                    for state, count in counts.most_common(12)
                ],
                "interpretation_boundary": "aggregate coded triad only; no row-level judgment, support adequacy, employer validity, worker capacity, source validity, or current guidance claim",
            }
        )

    return {
        "dataset_id": DATASET_ID,
        "profile_id": "2001_ABC_survey_A_B_C_linked_structured_triad_profile_v0_2026_05_22",
        "status": "linked_structured_triad_profile_unreviewed_no_text_no_promotion",
        "strict_A_B_C_pair_count": len(strict_pairs),
        "source_rows": {"A": len(a), "B_pairs": len(b), "C_pairs": len(c)},
        "profile_count": len(profiles),
        "profiles": profiles,
        "source_content_exported": False,
        "narrative_content_included": False,
        "row_level_ids_exported": False,
        "review_status": "unreviewed",
    }


def counts_total(counts: Counter[str], required: list[str], forbidden: list[str]) -> int:
    total = 0
    for state, count in counts.items():
        parts = set(state.split("|"))
        if all(item in parts for item in required) and not any(item in parts for item in forbidden):
            total += count
    return total


def write_markdown(data: dict[str, Any]) -> None:
    lines = [
        "# 2001 ABC Survey A/B/C Linked Structured Triad Profile",
        "",
        "作成日: 2026-05-22",
        "Lane: Falcon Lab",
        "状態: strict linked coded triad profile / no narrative text / 未レビュー / 昇格なし",
        "本文引用: なし",
        "",
        "## Position",
        "",
        "このprofileは、A票の事業所側負担/助言、B票の上司側支援状態、C票の本人側必要/有用性を、A/B/Cが同時に揃うペアで集計する。支援妥当性や正誤判断ではなく、三者紐付けデータがFalconの構造読みにどこまで厚みを足せるかを見るための coded triad layer である。",
        "",
        f"- strict A/B/C linked pairs: {data['strict_A_B_C_pair_count']}",
        "",
        "## Category Profiles",
        "",
        "| category | pairs | A burden/expected | B support | B needed-not-present | C need/useful | A burden + C need without B support | routes |",
        "|---|---:|---:|---:|---:|---:|---:|---|",
    ]
    for item in data["profiles"]:
        routes = ", ".join(f"`{route}`" for route in item["routes"])
        lines.append(
            f"| `{item['category']}` | {item['triad_pairs']} | "
            f"{item['A_burden_or_expected_burden']} ({item['A_burden_or_expected_burden_rate']}) | "
            f"{item['B_support_present']} ({item['B_support_present_rate']}) | "
            f"{item['B_supervisor_needed_but_not_present']} ({item['B_supervisor_needed_but_not_present_rate']}) | "
            f"{item['C_need_or_usefulness']} ({item['C_need_or_usefulness_rate']}) | "
            f"{item['A_burden_and_C_need_without_B_support']} | {routes} |"
        )

    lines.extend(["", "## Use Boundary", ""])
    lines.extend(
        [
            "- A burden is an establishment-side perception/expected-burden signal, not burden validity.",
            "- B support state is a supervisor/workplace-side report, not adequacy.",
            "- C need/usefulness is a worker-side signal, not capacity or entitlement judgment.",
            "- Counts are structural contrast windows only; human review is required before reusable knowledge movement.",
            "",
        ]
    )
    OUTPUT_MD.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    data = build_profile()
    OUTPUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_markdown(data)
    print(
        json.dumps(
            {
                "strict_A_B_C_pair_count": data["strict_A_B_C_pair_count"],
                "profile_count": data["profile_count"],
                "output": str(OUTPUT_JSON.relative_to(REPO_ROOT)),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
