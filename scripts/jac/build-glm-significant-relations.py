#!/usr/bin/env python3

from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
GLM_DIR = ROOT / "references" / "GLM_resutls"
INPUT_PATH = GLM_DIR / "NanbyoGLM.xlsx"
OUTPUT_RELATIONS_PATH = GLM_DIR / "nanbyo-glm-significant-relations.json"
OUTPUT_METRICS_PATH = GLM_DIR / "nanbyo-glm-metrics.json"

P_THRESHOLD = float(os.environ.get("GLM_P_THRESHOLD", "0.05"))

ROW_OUTCOME = 3
ROW_LABELS = 5
ROW_START = 7
COL_PREDICTOR_ID = 3
COL_PREDICTOR_GROUP = 4
COL_PREDICTOR_TEXT = 5

STOPWORDS = {
    "問題",
    "発生",
    "現状",
    "必要",
    "要因",
    "状況",
    "生活",
    "就業",
    "就労",
    "困難性",
    "解決可能性",
    "検討",
    "モデル",
    "支障",
    "機能障害等",
    "就職活動",
    "職業準備",
    "就業継続",
}


@dataclass(frozen=True)
class OutcomeColumn:
    pair_index: int
    p_col: int
    b_col: int
    outcome: str


def as_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("−", "-").replace("―", "-")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_text(value: Any) -> str:
    return str(value or "").replace("\r\n", "\n").replace("\r", "\n").strip()


def normalize_predictor_id(value: Any, row: int) -> str:
    num = as_float(value)
    if num is None:
        return f"row-{row}"
    if num.is_integer():
        return str(int(num))
    return str(num)


def extract_sheet_order(sheet_name: str) -> int:
    matched = re.match(r"^\s*(\d+)", sheet_name)
    if not matched:
        return 99
    return int(matched.group(1))


def build_relation_id(sheet_order: int, predictor_id: str, pair_index: int) -> str:
    safe_predictor = re.sub(r"[^0-9A-Za-z_-]", "-", predictor_id)
    return f"GLM-S{sheet_order}-P{safe_predictor}-C{pair_index:03d}"


def derive_keywords(*texts: str) -> list[str]:
    combined = " ".join([normalize_text(text) for text in texts if normalize_text(text)])
    if not combined:
        return []
    raw_parts = re.split(r"[\s\n\r\t,，.。/／・:：;；()（）\[\]「」『』]+", combined)
    keywords: list[str] = []
    seen = set()
    for part in raw_parts:
        token = part.strip()
        if len(token) < 2:
            continue
        if token in STOPWORDS:
            continue
        if token in seen:
            continue
        seen.add(token)
        keywords.append(token)
        if len(keywords) >= 14:
            break
    return keywords


def detect_outcome_columns(ws) -> list[OutcomeColumn]:
    outcomes: list[OutcomeColumn] = []
    pair_index = 0
    for p_col in range(1, ws.max_column):
        p_label = normalize_text(ws.cell(row=ROW_LABELS, column=p_col).value).lower()
        b_label = normalize_text(ws.cell(row=ROW_LABELS, column=p_col + 1).value).upper()
        if p_label != "p" or b_label != "B":
            continue
        pair_index += 1
        outcome = normalize_text(ws.cell(row=ROW_OUTCOME, column=p_col).value)
        outcomes.append(
            OutcomeColumn(
                pair_index=pair_index,
                p_col=p_col,
                b_col=p_col + 1,
                outcome=outcome,
            )
        )
    return outcomes


def relation_summary(predictor: str, outcome: str, b: float, p: float) -> str:
    return f"「{predictor}」と「{outcome}」に有意な関連（B={b:.3f}, p={p:.3g}）"


def main() -> None:
    wb = load_workbook(INPUT_PATH, data_only=True, read_only=False)
    generated_at = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")

    all_relations: list[dict[str, Any]] = []
    sheet_metrics: list[dict[str, Any]] = []

    for ws in wb.worksheets:
        sheet_name = ws.title.strip()
        if sheet_name == "Sheet1":
            continue

        sheet_order = extract_sheet_order(sheet_name)
        outcome_columns = detect_outcome_columns(ws)
        if not outcome_columns:
            continue

        significant_count = 0
        predictor_with_significant = set()

        for row in range(ROW_START, ws.max_row + 1):
            predictor = normalize_text(ws.cell(row=row, column=COL_PREDICTOR_TEXT).value)
            if not predictor:
                continue
            predictor_id = normalize_predictor_id(
                ws.cell(row=row, column=COL_PREDICTOR_ID).value, row=row
            )
            predictor_group = normalize_text(
                ws.cell(row=row, column=COL_PREDICTOR_GROUP).value
            )

            for outcome_col in outcome_columns:
                p_value = as_float(ws.cell(row=row, column=outcome_col.p_col).value)
                b_value = as_float(ws.cell(row=row, column=outcome_col.b_col).value)
                if p_value is None or b_value is None:
                    continue
                if p_value > P_THRESHOLD:
                    continue

                significant_count += 1
                predictor_with_significant.add(predictor_id)

                direction = "up" if b_value > 0 else "down"
                outcome = outcome_col.outcome or f"outcome_col_{outcome_col.p_col}"
                relation_id = build_relation_id(sheet_order, predictor_id, outcome_col.pair_index)
                all_relations.append(
                    {
                        "id": relation_id,
                        "sheet": sheet_name,
                        "sheetOrder": sheet_order,
                        "predictorId": predictor_id,
                        "predictorGroup": predictor_group,
                        "predictor": predictor,
                        "outcome": outcome,
                        "p": float(p_value),
                        "b": float(b_value),
                        "direction": direction,
                        "summary": relation_summary(predictor, outcome, b_value, p_value),
                        "keywords": derive_keywords(sheet_name, predictor_group, predictor, outcome),
                    }
                )

        sheet_metrics.append(
            {
                "sheet": sheet_name,
                "predictorRows": sum(
                    1
                    for row in range(ROW_START, ws.max_row + 1)
                    if normalize_text(ws.cell(row=row, column=COL_PREDICTOR_TEXT).value)
                ),
                "pColumns": len(outcome_columns),
                "significantRelations": significant_count,
                "predictorsWithSignificantRelation": len(predictor_with_significant),
            }
        )

    all_relations.sort(
        key=lambda item: (
            int(item["sheetOrder"]),
            float(item["p"]),
            -abs(float(item["b"])),
            str(item["predictorId"]),
            str(item["outcome"]),
        )
    )

    totals = {
        "significantRelations": len(all_relations),
        "uniquePredictorsWithSignificantRelation": len({item["predictor"] for item in all_relations}),
    }

    relations_payload = {
        "generatedAt": generated_at,
        "sourceFile": str(INPUT_PATH),
        "pThreshold": P_THRESHOLD,
        "relationCount": len(all_relations),
        "sheetCount": len(sheet_metrics),
        "relations": all_relations,
    }

    metrics_payload = {
        "generatedAt": generated_at,
        "sourceFile": str(INPUT_PATH),
        "sheetCount": len(sheet_metrics),
        "sheets": sheet_metrics,
        "totals": totals,
    }

    OUTPUT_RELATIONS_PATH.write_text(
        json.dumps(relations_payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    OUTPUT_METRICS_PATH.write_text(
        json.dumps(metrics_payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )

    print(
        json.dumps(
            {
                "generatedAt": generated_at,
                "relationCount": len(all_relations),
                "sheetCount": len(sheet_metrics),
                "totals": totals,
                "relationsPath": str(OUTPUT_RELATIONS_PATH),
                "metricsPath": str(OUTPUT_METRICS_PATH),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
